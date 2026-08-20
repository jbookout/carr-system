"""record_sources.py — one read path for the derived-surface consumers (ORDER 29a).

WHY THIS EXISTS
  The Deal Room and the Obsidian graph were built when the generated files WERE
  the record: they opened vendors.xlsx / lead-registry.xlsx / client-roster.xlsx /
  panhandle-team-deals.json and derived their surfaces from whatever those files
  happened to hold. Since the record layer went live those files are exports —
  the database is the record and the file is a rendering of it. A local consumer
  that reads the rendering is one export failure away from deriving a board from
  yesterday's truth without noticing.

  ORDER 28's inventory found the same thing the other way round: Cowork sessions
  and Dell's side have no database path at all, so the files are permanent
  infrastructure, not scar tissue. Nothing is retired here. The files keep being
  generated exactly as before; only LOCAL CODE gains a second way to read them.

WHAT IT GUARANTEES
  Byte-identical derived output in either mode. The record path queries the same
  views the exporters query (`v_export_vendors`, `v_export_leads`,
  `v_export_clients`, `v_export_deals`), projects the same columns in the same
  order, applies the same filters, and reproduces the two shape changes a
  round-trip through the file makes:
    1. xlsx has no date type — openpyxl hands back `datetime` at midnight where
       the view hands back `date`. Records mode promotes date -> datetime so a
       consumer sees the shape it has always seen.
    2. JSON has no date or Decimal type — the deals exporter serialises with
       `default=str`. Records mode round-trips the deal list through json for
       the same reason.
  Proven, not asserted: `tools/parity-records.py` runs every consumer in both
  modes on the same day's data and diffs the derived output.

MODE SELECTION (per consumer, highest precedence first)
    --files / --records on the command line
    CARR_SOURCE_MODE=files|records in the environment
    the consumer's own default
  Records mode falls back to files, loudly on stderr, when there is no exporter
  credential or psycopg is not importable (a plain `python3`, Dell's runtime, a
  machine with no db.env). A credential that is present but fails is NOT
  swallowed: that is a real outage and it raises.

The column lists come from `exporters/targets.py` by import, never by copy, so
they cannot drift from what the exporters actually write.
"""

import json
import os
import sys
from datetime import date, datetime
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent

MODE_FILES = "files"
MODE_RECORDS = "records"

VENDORS_REL = "DNA/Network/vendors.xlsx"
LEADS_REL = "DNA/Leads/lead-registry.xlsx"
CLIENTS_REL = "DNA/Clients/client-roster.xlsx"
DEALS_REL = "DNA/Deal Management/panhandle-team-deals.json"

# ---------------- the prospect pool (ORDER 26) ----------------
#
# The pool is a different read from the four above, and the reason is a real
# boundary rather than an inconvenience. ORDER 25 deliberately gave the pool two
# narrow surfaces:
#   v_pool         — safe columns, granted to carr_reader. No address, no email,
#                    no phone, no source_row. 9,320 uncontacted third parties is
#                    an order of magnitude more personal data than the 207 worked
#                    leads, so the class-parity argument that opened v_export_leads
#                    does not carry here.
#   v_export_pool  — full columns INCLUDING source_row, granted to carr_exporter
#                    only, and scoped `where source = 'lead-router'` because it
#                    exists to regenerate the router sheet (export target #8).
#
# The board needs full columns for EVERY source, which is neither of those. The
# read therefore prefers an elevated DSN (what `tools/db-tap.py run` sets) and
# falls back to the exporter's router-only slice. When the board asks for lane
# sources and only the router slice is reachable, records mode is UNAVAILABLE and
# the caller falls back to files loudly — it never renders a board that is
# quietly missing 540 radar rows. Closing that gap properly is one line in a
# migration (an all-source export view, or a select grant), which ORDER 26 is not
# allowed to write; it is parked in the execution log instead.

ROUTER_SOURCE = "lead-router"
LANE_SOURCES = ("corp-filings", "upstream", "renewal-radar",
                "relocating-owner", "national-accounts")


# ---------------- mode selection ----------------

def resolve_mode(argv, default=MODE_RECORDS):
    """(mode, argv-without-the-mode-flags). Pass sys.argv[1:]; positional args survive."""
    mode = os.environ.get("CARR_SOURCE_MODE") or default
    rest = []
    for a in argv:
        if a == "--files":
            mode = MODE_FILES
        elif a == "--records":
            mode = MODE_RECORDS
        else:
            rest.append(a)
    if mode not in (MODE_FILES, MODE_RECORDS):
        sys.exit(f"unknown source mode {mode!r} (files|records)")
    return mode, rest


def _exporter_url():
    """Same lookup exporters/common.py does. Returns None when unconfigured."""
    url = os.environ.get("CARR_DB_EXPORTER_URL")
    if url:
        return url
    env = Path.home() / ".config/carr/db.env"
    if env.exists():
        for line in env.read_text().splitlines():
            if line.startswith("CARR_DB_EXPORTER_URL="):
                # .strip("\"'") — db.env values are shell-quoted so `set -a; . db.env`
                # survives an `&` in the DSN; psycopg needs them unquoted. Full reasoning
                # in exporters/common.py. Added 2026-08-02.
                return line.split("=", 1)[1].strip().strip("\"'") or None
    return None


def _records_available():
    """(bool, why-not). Missing credential or missing driver is a fallback, not a failure."""
    if not _exporter_url():
        return False, "no CARR_DB_EXPORTER_URL (see ~/.config/carr/db.env)"
    try:
        import psycopg  # noqa: F401
    except ImportError:
        return False, "psycopg not importable by this interpreter (use .venv/bin/python)"
    return True, ""


def effective_mode(mode, label="source"):
    """Resolve records -> files when the record path is unreachable. Says so on stderr."""
    if mode != MODE_RECORDS:
        return MODE_FILES
    ok, why = _records_available()
    if ok:
        return MODE_RECORDS
    print(f"[{label}] records mode unavailable ({why}) — falling back to the generated files",
          file=sys.stderr)
    return MODE_FILES


def source_note(mode):
    return "records (v_export_* views)" if mode == MODE_RECORDS else "generated files"


# ---------------- shared shape rules ----------------

def _as_file_shape(v):
    """A date read back out of an xlsx is a datetime at midnight. Match it."""
    if type(v) is date:
        return datetime(v.year, v.month, v.day)
    return v


def _connect():
    sys.path.insert(0, str(REPO))
    from exporters.common import connect
    return connect()


def _view_rows(query, colspec, drop_when=None):
    """Rows of `query` projected onto `colspec`, in view order, file-shaped."""
    with _connect() as conn, conn.cursor() as cur:
        cur.execute(query)
        cols = [d[0] for d in cur.description]
        data = cur.fetchall()
    out = []
    for r in data:
        if drop_when is not None and r[cols.index(drop_when)]:
            continue
        out.append({c: _as_file_shape(r[cols.index(c)]) for c in colspec})
    return out


def _sheet_rows(path, sheet):
    """The historical read: one dict per non-empty data row, keyed by header."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h).strip() if h else "" for h in next(it, [])]
    out = []
    for r in it:
        d = {hdr[i]: (r[i] if i < len(r) else None) for i in range(len(hdr))}
        if any(v not in (None, "") for v in d.values()):
            out.append(d)
    wb.close()
    return out


def _targets():
    sys.path.insert(0, str(REPO))
    from exporters import targets
    return targets


# ---------------- the four sources ----------------

def load_vendors(root, mode):
    if mode == MODE_RECORDS:
        t = _targets()
        # `_out_of_market` rows never reach the Vendors sheet, so they must not
        # reach a records-mode consumer either.
        return _view_rows('select * from v_export_vendors order by "ID"',
                          t.VENDORS_COLS, drop_when="_out_of_market")
    return _sheet_rows(os.path.join(root, VENDORS_REL), "Vendors")


def load_leads(root, mode):
    if mode == MODE_RECORDS:
        t = _targets()
        return _view_rows('select * from v_export_leads order by "Lead ID"', t.REGISTRY_COLS)
    return _sheet_rows(os.path.join(root, LEADS_REL), "Registry")


def load_clients(root, mode):
    if mode == MODE_RECORDS:
        t = _targets()
        return _view_rows('select * from v_export_clients order by "Client ID"', t.ROSTER_COLS)
    return _sheet_rows(os.path.join(root, CLIENTS_REL), "Clients")


def load_deals_doc(root, mode):
    """The whole deal document: {'deals': [...], 'captured': ...} plus the file's prose.

    The merge below mirrors `exporters/targets.py:build_deals` — the fidelity rule
    (legacy passthrough wins for txn/seg/carr, the DB wins for name/phase/owner)
    is the exporter's, not this module's. targets.py is the authority; parity is
    what proves this copy still agrees with it.
    """
    if mode != MODE_RECORDS:
        path = os.path.join(root, DEALS_REL)
        if not os.path.exists(path):
            return {"deals": [], "captured": ""}
        with open(path) as fh:
            return json.load(fh)

    with _connect() as conn, conn.cursor() as cur:
        cur.execute("select * from v_export_deals order by name")
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
    deals = []
    for r in rows:
        row = dict(zip(cols, r))
        legacy = row.get("source_row") or {}
        legacy.update({
            "name": row["name"], "phase": row["phase"],
            "owner": row["owner"] or legacy.get("owner"),
            "txn": legacy["txn"] if "txn" in legacy else row["deal_type"],
            "seg": legacy["seg"] if "seg" in legacy else row["segment"],
            "carr": legacy.get("carr") if legacy.get("carr") is not None
                    else row["PLACEHOLDER_sf_commission_never_sum"],
        })
        # 0074: city/lane are real columns; the DB owns them. Mirrors
        # exporters/targets.py:build_deals exactly — targets.py is the authority
        # and this block must not drift from it. Correct on both sides of the
        # migration (column absent -> legacy passthrough still answers).
        for _f in ("city", "lane"):
            if row.get(_f) is not None:
                legacy[_f] = row[_f]
        deals.append(legacy)
    # The file goes through json.dumps(default=str); a records-mode consumer must
    # see the same scalars (no date, no Decimal) or its output differs on type alone.
    deals = json.loads(json.dumps(deals, default=str))
    from datetime import timezone
    return {
        "source": "READ LIVE from the CARR record layer (v_export_deals)",
        "captured": datetime.now(timezone.utc).isoformat(),
        "count": len(deals),
        "deals": deals,
    }


# ---------------- the intro graph (ORDER 32) ----------------

def load_party_links(mode):
    """The intro graph's real edges, or None when only the files are reachable.

    RECORDS ONLY, AND THE None IS THE POINT. `party_link` has no file twin: the
    xlsx `Links` column is the PROSE the edges were parsed out of, not the edges,
    and re-deriving edges from that prose inside the graph pipeline would be a
    second parser with a second grammar quietly disagreeing with
    `parse_party_links.py`. Two vocabularies is the mistake ORDER 18 already had
    to unpick once. So a files-mode consumer gets None and keeps its own
    historical behaviour, out loud, rather than a half-graph that looks complete.

    Reads `v_party_graph` — the ORDER 18 reader-safe view: refs, names, the kind
    and the provenance note, never contact detail — so the graph pipeline sees
    exactly what a reader-scoped session sees.

    EVERY EDGE COMES BACK NOW, INCLUDING THE ONES WITH A NULL REF (loop #133,
    2026-08-02). This used to filter `from_ref is not null and to_ref is not
    null` in SQL, which meant seven of thirty-one edges never reached the
    consumer at all — the pipeline reported "31 edges" only because it happened
    to be told 24, and six of the seven were Joe's own `can_introduce` edges,
    the highest-value class in the referral engine. A filter in the loader is a
    filter nobody downstream can see or count. The consumer now decides what to
    do with a null endpoint (build-graph-notes.py resolves it by exact name and
    reports whatever it still cannot map), which is the same posture the rest of
    this module takes: hand over the record, do not quietly shrink it.
    """
    if mode != MODE_RECORDS:
        return None
    with _connect() as conn, conn.cursor() as cur:
        cur.execute("""select from_ref, from_name, kind, to_ref, to_name, note
                         from v_party_graph
                        order by from_ref nulls last, kind, to_ref nulls last""")
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def load_ref_index(mode):
    """Every ref the record layer carries — {ref, name, kind, merged, party_id} — or None.

    WHY A CONSUMER NEEDS THIS (loop #133). `v_party_graph` resolves an edge
    endpoint to ONE business ref (C- / V- / L-) and hands back NULL when the
    party has none. Two real cases produce that null and they need opposite
    treatment, so a consumer has to be able to look the party up:

      · a BARE PARTY — Joe Bookout is P-1084 with no client, lead or vendor row,
        because he is the agent, not a record in his own pipeline. He is a
        legitimate node and the graph should draw him.
      · a TOMBSTONE the link still points at — party_link 4aecf3b0 points at
        P-0365, merged into P-0384 (one real lead). The survivor carries
        C-155 and L-208; the loser carries no role at all, hence the null. That
        edge should resolve to the SURVIVOR, not become a second node for the
        same person. (Example sanitized 2026-08-06, ORDER 42b — the original
        named the real lead.)

    `v_ref_index` distinguishes the two: the party branch carries P- refs and
    the merged flag, so a live bare party is findable and a tombstone is
    identifiable. Reader-safe columns only — name, ref, kind, merged, party_id —
    the same posture load_party_links takes.

    party_id comes back as a string so it can be a dict key without a uuid
    import leaking into every consumer.
    """
    if mode != MODE_RECORDS:
        return None
    with _connect() as conn, conn.cursor() as cur:
        cur.execute("""select ref, display_name, subject_type, merged, party_id
                         from v_ref_index
                        where ref is not null
                        order by subject_type, ref""")
        return [{"ref": r[0], "name": r[1], "kind": r[2],
                 "merged": bool(r[3]), "party_id": str(r[4]) if r[4] else None}
                for r in cur.fetchall()]


# ---------------- the prospect pool ----------------

def _elevated_url():
    """A DSN that can read candidate_pool itself. `tools/db-tap.py run` sets DATABASE_URL."""
    # The scheduled-jobs role already has the narrow pool grant needed to map
    # lane output.  Let that same non-owner role read the canonical rows during
    # the run, instead of making an unattended routine fall back to Drive simply
    # because it does not carry an exporter credential.
    return (os.environ.get("CARR_DB_POOL_URL")
            or os.environ.get("CARR_DB_JOBS_URL")
            or os.environ.get("DATABASE_URL"))


# Which relation a reachable pool read comes from. Returned by pool_reach as its
# fourth element and branched on by load_pool. All three project the same keys.
POOL_BASE = "candidate_pool"        # elevated DSN, base table, every source
POOL_ALL = "v_export_pool_all"      # exporter credential, every source (0025)
POOL_ROUTER = "v_export_pool"       # exporter credential, source='lead-router' only


def pool_reach(wanted):
    """(ok, why-not, connect-fn, relation) for the pool sources `wanted` needs.

    Reports rather than raises: the board turns a miss into a loud fallback to
    files, which is the ORDER 29a contract for an unreachable record path.

    The exporter credential reaches every source through `v_export_pool_all`
    (migration 0025, ORDER 26's parked flip). `v_export_pool` stays router-scoped
    because it IS export target #8 and must not grow the lane rows; the all-source
    view is a consumer read path and is not an export target. This function
    refused the lanes outright until 2026-08-04 — the view had been applied since
    2026-07-31 and nothing here had been taught to look for it.
    """
    try:
        import psycopg  # noqa: F401
    except ImportError:
        return False, "psycopg not importable by this interpreter (use .venv/bin/python)", None, None
    url = _elevated_url()
    if url:
        def _c():
            import psycopg
            return psycopg.connect(url)
        try:
            with _c() as conn, conn.cursor() as cur:
                cur.execute("select 1 from candidate_pool limit 1")
        except Exception as e:                                  # noqa: BLE001
            return False, f"elevated DSN cannot read candidate_pool ({type(e).__name__})", None, None
        return True, "", _c, POOL_BASE
    if not _exporter_url():
        return False, "no CARR_DB_POOL_URL and no CARR_DB_EXPORTER_URL", None, None
    if not set(wanted) - {ROUTER_SOURCE}:
        return True, "", _connect, POOL_ROUTER
    # Lanes wanted. Prove the all-source view rather than assuming the grant:
    # a missing view here must fall back loudly, not raise mid-board.
    try:
        with _connect() as conn, conn.cursor() as cur:
            cur.execute(f"select 1 from {POOL_ALL} limit 1")
    except Exception as e:                                      # noqa: BLE001
        return (False,
                f"the exporter credential cannot read {POOL_ALL} ({type(e).__name__}), so "
                f"{sorted(set(wanted) - {ROUTER_SOURCE})} are unreachable; "
                "migration 0025 creates and grants it",
                None, None)
    return True, "", _connect, POOL_ALL


def load_pool(wanted):
    """{source: [source_row dict, ...]} in each source's own file order.

    source_row is the finder's row VERBATIM, which is what makes a records-mode
    consumer able to reproduce a file-mode one exactly: the transform is shared,
    only where the raw row came from changes.
    """
    ok, why, conn_fn, relation = pool_reach(wanted)
    if not ok:
        raise RuntimeError(why)

    # THE ROUTER ROWS GET THE EXPORT'S OWN PROJECTION, and this is not cosmetic.
    # A router row exists twice inside the pool: nine sheet columns are DB-OWNED
    # (they have their own typed column, normalised by the importer's val()), and
    # source_row holds the sheet cell VERBATIM. exporters/targets.py:build_router
    # renders the DB-owned nine from the columns and everything else from
    # source_row, so the router xlsx a file-mode board reads carries the column
    # values. Measured 2026-07-31: 923 rows differ between the two — the empty
    # phone placeholder is '() ' verbatim and '()' after val() — and the board
    # would have shown one in records mode and the other in file mode.
    #
    # So this applies build_router's projection, importing the mapping from
    # targets.py rather than copying it. The lanes deliberately do NOT get the
    # same treatment: their board rows carry lane-specific fields that have no
    # column at all (le / tier / conf / ll / rep / newll / in_territory /
    # sensitivity), the relocating-owner lane is appended to the board WHOLESALE,
    # and overlaying columns onto those objects would inject changes where no
    # divergence exists.
    sys.path.insert(0, str(REPO))
    from exporters.targets import ROUTER_DB_OWNED

    sel = ", ".join(f'{v} as "{k}"' for k, v in _ROUTER_DB_COLS.items())
    cols = ", ".join(f'"{k}"' for k in ROUTER_DB_OWNED)
    if relation == POOL_BASE:
        q = (f"select source, source_seq, source_row, {sel} from {POOL_BASE} "
             "where source = any(%s) order by source, source_seq")
        args = (list(wanted),)
    elif relation == POOL_ALL:
        # The view already publishes v_export_pool's aliases, so the projection is
        # the router branch's, widened by `source` and filtered to what was asked.
        q = (f"select source, source_seq, source_row, {cols} from {POOL_ALL} "
             "where source = any(%s) order by source, source_seq")
        args = (list(wanted),)
    else:
        q = (f"select %s::text as source, source_seq, source_row, {cols} "
             f"from {POOL_ROUTER} order by source_seq")
        args = (ROUTER_SOURCE,)

    out = {s: [] for s in wanted}
    with conn_fn() as conn, conn.cursor() as cur:
        cur.execute(q, args)
        names = [d[0] for d in cur.description]
        for r in cur.fetchall():
            rec = dict(zip(names, r))
            src = rec["source"]
            if src not in out:
                continue
            row = dict(rec["source_row"] or {})
            if src == ROUTER_SOURCE:
                for sheet_col in ROUTER_DB_OWNED:
                    row[sheet_col] = rec[sheet_col]
            out[src].append(row)
    return out


# view alias -> base column, for the all-source read. The aliases are
# v_export_pool's, so both read paths hand the caller the same keys.
_ROUTER_DB_COLS = {
    "SEGMENT": "segment", "THE PLAY": "segment_play", "Name": "name",
    "Profession": "vertical", "Practice Address": "address", "City": "city",
    "County": "county", "Email": "email", "Phone": "phone",
}


# ---------------- doctrine store: the vault's third-and-fourth readers (Phase 1) ----------------
#
# retrieve.py already ran a dual-read pass against doctrine_migration_batch to
# avoid printing a stale file pointer for content the store now owns. Phase 1
# (2026-08-13, the Aug 21 retirement gate) gives build-section-index.py and
# build-system-graph.py the SAME two facts, so all three systemic vault readers
# compute "is this file store-held now" and "what does the store actually say"
# from ONE function each — a second implementation of either is exactly the
# drift rule 73381d78 exists to prevent. carr_exporter already holds SELECT on
# every doctrine_* table read here (0078_exporter_doctrine_select.sql).
#
# doctrine_document carries no source_path column (P1 schema, 0075) — imports
# only ever recorded the ORIGINAL FILE LIST on the batch ledger. doctrine_slug_by_path
# reconstructs the slug doctrine_import.py would have assigned to each migrated
# path, replaying the same kebab-case + index/readme-parent-prefix + collision
# rule against doctrine_migration_batch rows in (batch_no, source_paths order) —
# the same order the batches were actually applied in. Verified 2026-08-13
# against the live store: 217/217 reconstructed slugs match a real document row.
# If a future batch breaks that (a collision rule change, a hand-edited slug),
# the mismatch is silently DROPPED rather than mis-attributed — a graph edge
# that cannot be placed is a smaller loss than one placed on the wrong folder.

def _strip_source_root(path, source_root):
    """Return a record source identity relative to its declared source root.

    This deliberately has no machine-specific mount aliases.  Callers in the
    normal path pass their repository root, and a recovery caller must pass the
    root it explicitly selected.  Ambient mount configuration is not evidence
    of a canonical source identity.
    """
    root = os.path.abspath(os.fspath(source_root))
    candidate = os.path.abspath(os.fspath(path))
    try:
        return os.path.relpath(candidate, root) if os.path.commonpath((root, candidate)) == root else path
    except ValueError:
        return path
    return path


def doctrine_migrated_paths(source_root):
    """Vault-relative paths recorded on a VERIFIED migration batch. Content for
    these lives in the store now — a file-walking consumer must stop opening
    them, exactly what retrieve.py's dual-read pass already assumes."""
    migrated = set()
    with _connect() as conn, conn.cursor() as cur:
        cur.execute("select source_paths from doctrine_migration_batch where state = 'verified'")
        for (paths,) in cur.fetchall():
            for p in paths or []:
                migrated.add(_strip_source_root(p, source_root))
    return migrated


def doctrine_slug_by_path(source_root):
    """{vault-relative path: document slug} for every migrated file, replaying
    doctrine_import.py's kebab/collision rule against the batch ledger (no file
    opens — the ledger alone is enough to reproduce it, see module note above)."""
    import re
    import unicodedata

    def kebab(s):
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
        s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
        return s or "section"

    mapping = {}
    assigned = set()
    with _connect() as conn, conn.cursor() as cur:
        cur.execute("select source_paths from doctrine_migration_batch "
                     "where state = 'verified' order by batch_no")
        for (paths,) in cur.fetchall():
            for p in paths or []:
                stem = Path(p).stem
                parent = Path(p).parent.name
                if stem.lower() in ("index", "readme"):
                    slug = kebab(f"{parent}-{stem}") if parent and parent != "CARR AI" \
                        else kebab(f"root-{stem}")
                else:
                    slug = kebab(stem)
                if slug in assigned:
                    slug = f"{kebab(parent)}-{slug}"
                assigned.add(slug)
                mapping[_strip_source_root(p, source_root)] = slug
    return mapping


def doctrine_sections(visibility="shared"):
    """Every active doctrine section, store identity attached — one read-only
    query. [{slug, doc_title, content_class, section_key, title, ordinal,
    plain_text}, ...], ordered by document then section position. Default
    visibility='shared' matches retrieve.py's own FTS pass; every document is
    'shared' as of this build (no 'personal' doctrine exists yet)."""
    with _connect() as conn, conn.cursor() as cur:
        cur.execute("""
            select d.slug, d.title, d.content_class, s.section_key, s.title,
                   s.ordinal, r.plain_text
              from doctrine_section s
              join doctrine_document d on d.id = s.document_id
              join doctrine_revision r on r.id = s.current_revision_id
             where s.status = 'active' and d.visibility = %s
             order by d.slug, s.ordinal""", (visibility,))
        cols = ["slug", "doc_title", "content_class", "section_key",
                "title", "ordinal", "plain_text"]
        return [dict(zip(cols, row)) for row in cur.fetchall()]
