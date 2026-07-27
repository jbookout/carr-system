#!/usr/bin/env python3
"""
registry.py — the ONE sanctioned writer for lead-registry.xlsx.

WHY THIS EXISTS (built 2026-07-27, after the 7/25 sweep intake)
  `sheets.py` (orchestrator corrective #1, 2026-07-25) guards the READ side: no
  consumer may read the registry positionally. Nothing guarded the WRITE side,
  because no repo script writes the registry — every write to date was ad-hoc
  session code. Two defects came straight out of that gap:

  #100  The 7/25 write rebuilt the Registry sheet with 23 columns instead of 26.
        Est-Lease-Event, Event-Source and Event-Confidence (lane 3) were dropped
        silently, taking every lease-event datapoint with them.

  #101  The same write allocated L-164 … L-200 on top of three LIVE rows:
        L-164 First Call DPC (Hot – Startup, ETL signed, Salesforce deal open),
        L-165 Amanda Troupe / Troupe Vision of Florida PLLC,
        L-166 Coast Performance Medicine LLC.
        Its Intake Log line reads "0 dedup collisions vs registry" — true, and
        irrelevant: it deduped on name, and allocated from the INTAKE LOG's
        last-recorded ID (L-163, the 7/13 Renalus line) instead of the Registry
        sheet's real max (L-166). The 7/14 and 7/16 adds were never ledgered, so
        the ledger was three IDs stale and the writer trusted it.

  Both are now structurally impossible through this module:
    - allocate_ids() reads the REGISTRY SHEET, never the ledger, and never
      reissues a retired ID (L-162 was retired by design on 2026-07-09; the
      Intake Log says so — it is a tombstone, not a casualty).
    - write_guarded() refuses to save if any pre-existing row lost its occupant,
      disappeared, or if any canonical column went missing.

RULE (lead-system.md): no session writes lead-registry.xlsx by hand. Bulk intake
goes through append_rows(); a field edit on an existing row goes through
edit_row(). Both back up first and both run the invariants on the way out.
"""
import os
import re
import shutil
import datetime

import openpyxl

# The canonical Registry vocabulary, in order. A write that cannot produce every
# one of these halts. Lane 3 (the last three) is the reason this list is here.
REGISTRY_COLUMNS = [
    "Lead ID", "Date In", "Owner", "Stage", "Segment", "Contact Name",
    "Practice", "Specialty", "City/Market", "County", "Email", "Phone",
    "Source Type", "Source Detail (V-ID / event / referrer)", "Report-Back Due",
    "Drip Campaign", "Drip Added", "Next Action", "Next Action Date",
    "Last Touch", "SF Deal", "Detail File", "Notes",
    "Est-Lease-Event", "Event-Source", "Event-Confidence",
]

# IDs deliberately taken out of service. Never reissued, never reported as loss.
#   L-162 — Dell's joint-test row, 2026-07-09; deleted same day after Joe
#           confirmed T6 (twin live). Intake Log: "L-162 stays retired".
RETIRED_IDS = {162}

ID_RE = re.compile(r"^L-0*(\d+)$")


def _s(v):
    return str(v if v is not None else "").strip()


def id_num(v):
    """'L-164' -> 164. None for anything that is not a well-formed Lead ID."""
    m = ID_RE.match(_s(v))
    return int(m.group(1)) if m else None


def fmt_id(n):
    return "L-%03d" % n


# ---------------------------------------------------------------- read side

def open_registry(path):
    """Load the workbook and return (wb, ws, colmap). Halts on a missing column."""
    wb = openpyxl.load_workbook(path)
    if "Registry" not in wb.sheetnames:
        raise SystemExit("SCHEMA HALT [%s]: no 'Registry' sheet." % path)
    ws = wb["Registry"]
    colmap = assert_columns(ws, path)
    return wb, ws, colmap


def assert_columns(ws, label):
    """{header: 1-based column index}. Dies loudly if a canonical column is gone.

    This is the #100 guard. It fires on the way IN as well as on the way out, so
    a workbook that has already lost a column cannot be quietly appended to.
    """
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    m = {_s(h): i + 1 for i, h in enumerate(hdr) if _s(h)}
    missing = [c for c in REGISTRY_COLUMNS if c not in m]
    if missing:
        raise SystemExit(
            "SCHEMA HALT [%s]: Registry is missing canonical column(s): %s. "
            "Present: %s. Restore the columns from the newest backup before writing — "
            "do NOT append around a missing column (that is defect #100)."
            % (label, missing, sorted(m))
        )
    return m


def snapshot(ws, colmap):
    """{id_num: (Contact Name, Practice)} for every data row. The occupant map.

    write_guarded diffs against this. Any pre-existing ID whose occupant changes
    is the #101 signature, whatever the caller thought it was doing.
    """
    out = {}
    dupes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(x is not None and _s(x) for x in row):
            continue
        n = id_num(row[colmap["Lead ID"] - 1])
        if n is None:
            continue
        occ = (_s(row[colmap["Contact Name"] - 1]), _s(row[colmap["Practice"] - 1]))
        if n in out:
            dupes.append(n)
        out[n] = occ
    if dupes:
        raise SystemExit(
            "INTEGRITY HALT: duplicate Lead ID(s) already present: %s. "
            "Resolve before writing." % sorted(set(dupes))
        )
    return out


def allocate_ids(ws, colmap, count):
    """The next `count` free IDs, taken from the SHEET's max and going up.

    Never reuses a gap and never reuses a retired ID. Gaps below the max are
    tombstones (a row that was deliberately removed); reissuing one would point
    old references at a new person, which is the same class of harm as #101.
    """
    used = set(snapshot(ws, colmap)) | RETIRED_IDS
    n = max(used)
    out = []
    while len(out) < count:
        n += 1
        out.append(fmt_id(n))
    return out


# --------------------------------------------------------------- write side

def backup(path, tag):
    """Timestamped copy beside the original. Returns the backup path."""
    today = datetime.date.today().isoformat()
    base, ext = os.path.splitext(path)
    dest = "%s.backup-%s-%s%s" % (base, today, tag, ext)
    i = 2
    while os.path.exists(dest):
        dest = "%s.backup-%s-%s-%d%s" % (base, today, tag, i, ext)
        i += 1
    shutil.copy2(path, dest)
    return dest


def write_guarded(path, wb, ws, before, tag):
    """Run the invariants, back up, then save. The only sanctioned save.

    `before` is the snapshot() taken before the caller touched anything.
    Refuses to save on any of:
      - a canonical column missing            (#100)
      - a pre-existing ID that lost its row   (#101)
      - a pre-existing ID whose occupant changed (#101)
      - a duplicate ID
      - a retired ID brought back to life
    """
    assert_columns(ws, path)
    after = snapshot(ws, colmap=assert_columns(ws, path))

    vanished = sorted(set(before) - set(after))
    if vanished:
        raise SystemExit(
            "INTEGRITY HALT: %d pre-existing row(s) would disappear: %s. "
            "A write may add and may edit; it may not delete. Nothing was saved."
            % (len(vanished), [fmt_id(n) for n in vanished])
        )

    stolen = [n for n in before if n in after and after[n] != before[n]]
    if stolen:
        detail = "; ".join(
            "%s was %s / %s, would become %s / %s"
            % (fmt_id(n), before[n][0] or "(blank)", before[n][1] or "(blank)",
               after[n][0] or "(blank)", after[n][1] or "(blank)")
            for n in sorted(stolen)
        )
        raise SystemExit(
            "INTEGRITY HALT: %d existing Lead ID(s) would change occupant. %s. "
            "This is defect #101. Allocate new IDs with allocate_ids() instead. "
            "Nothing was saved." % (len(stolen), detail)
        )

    revived = sorted(set(after) & RETIRED_IDS)
    if revived:
        raise SystemExit(
            "INTEGRITY HALT: retired ID(s) reissued: %s. Retired IDs stay retired. "
            "Nothing was saved." % [fmt_id(n) for n in revived]
        )

    bpath = backup(path, tag)
    wb.save(path)
    return bpath, len(after) - len(before)


def log_intake(wb, run_date, source, rows_found, ids_added, notes):
    """Append the Intake Log line. Not optional.

    The 7/14 and 7/16 adds skipped this, the ledger went stale, and the 7/25
    writer allocated off the stale ledger. Every write that adds a row logs it.
    """
    ws = wb["Intake Log"]
    ws.append([run_date, source, rows_found, ids_added, notes])


def append_rows(path, records, source, notes, tag="intake"):
    """Add new leads. IDs are allocated here; a caller-supplied Lead ID is ignored.

    `records` is a list of dicts keyed by REGISTRY_COLUMNS names. Returns the
    list of assigned IDs.
    """
    wb, ws, colmap = open_registry(path)
    before = snapshot(ws, colmap)
    ids = allocate_ids(ws, colmap, len(records))

    for rec, lid in zip(records, ids):
        unknown = [k for k in rec if k not in colmap]
        if unknown:
            raise SystemExit(
                "SCHEMA HALT [%s]: record has non-canonical field(s) %s. "
                "Nothing was saved." % (path, unknown)
            )
        row = [None] * len(colmap)
        for k, v in rec.items():
            row[colmap[k] - 1] = v
        row[colmap["Lead ID"] - 1] = lid
        ws.append(row)

    span = ids[0] if len(ids) == 1 else "%s … %s" % (ids[0], ids[-1])
    log_intake(wb, datetime.date.today().isoformat(), source, len(records), span, notes)
    bpath, delta = write_guarded(path, wb, ws, before, tag)
    return ids, bpath, delta


if __name__ == "__main__":
    print(__doc__)
