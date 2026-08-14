#!/usr/bin/env python3
"""Match calendar attendee emails to people already in the record.

Reads the LOCAL Apple Calendar database (Full Disk Access required; EventKit is
not usable here because macOS cannot prompt an unbundled binary), pulls attendee
email addresses, and matches them against the client roster and lead registry.

Produces INFERRED TOUCHES: dated evidence that contact happened, each carrying the
event that proves it and a confidence level. Nothing is written to the record —
these are proposals for a partner to confirm, per the standing rule that the
system infers contact from evidence it already holds and never asks the partner to
report his own activity.

Match tiers, strongest first:
  exact   — attendee email equals a contact email in the record
  domain  — attendee's domain matches the domain of a known contact's email
  none    — external address with no counterpart in the record (a research lead)

Internal @carr.us addresses are reported separately and never counted as client
contact.
"""

import json
import os
import shutil
import sqlite3
import sys
import tempfile
import time
from collections import defaultdict

APPLE_EPOCH = 978307200
DEFAULT_DAYS = 120
GROUP_CONTAINER = os.path.expanduser(
    "~/Library/Group Containers/group.com.apple.calendar/Calendar.sqlitedb"
)
EXPORTS = os.path.expanduser("~/carr-system/out/exports")
INTERNAL_DOMAIN = "carr.us"
FREEMAIL = {"gmail.com", "icloud.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com"}


def load_record_contacts():
    """Return (email -> label) and (domain -> label) from the roster and registry."""
    import openpyxl

    by_email, by_domain = {}, {}

    def ingest(path, sheet, id_col, name_col, org_col, email_col):
        full = os.path.join(EXPORTS, path)
        if not os.path.exists(full):
            return
        wb = openpyxl.load_workbook(full, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            return
        ws = wb[sheet]
        header, idx = None, {}
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if header is None:
                if any(c.lower() == "email" for c in cells):
                    header = cells
                    for want, key in ((id_col, "id"), (name_col, "name"),
                                      (org_col, "org"), (email_col, "email")):
                        for i, c in enumerate(cells):
                            if c.lower() == want.lower():
                                idx[key] = i
                continue
            if "email" not in idx:
                continue
            email = cells[idx["email"]].strip().lower() if idx["email"] < len(cells) else ""
            name = cells[idx.get("name", 0)] if idx.get("name", 0) < len(cells) else ""
            org = cells[idx.get("org", 0)] if idx.get("org", 0) < len(cells) else ""
            ref = cells[idx.get("id", 0)] if idx.get("id", 0) < len(cells) else ""
            label = " / ".join(x for x in (ref, name or org) if x) or "(unnamed row)"
            if email and "@" in email:
                by_email.setdefault(email, label)
                dom = email.split("@", 1)[1]
                if dom not in FREEMAIL and dom != INTERNAL_DOMAIN:
                    by_domain.setdefault(dom, label)

    ingest("client-roster.xlsx", "Clients", "Client ID", "Name", "Practice / Entity", "Email")
    ingest("lead-registry.xlsx", "Registry", "Lead ID", "Contact Name", "Practice", "Email")
    return by_email, by_domain


def read_calendar(days):
    if not os.path.exists(GROUP_CONTAINER):
        print(f"FATAL: calendar database not found at {GROUP_CONTAINER}")
        return None
    tmp = tempfile.mkdtemp(prefix="calmatch-")
    local = os.path.join(tmp, "Calendar.sqlitedb")
    try:
        for suffix in ("", "-wal", "-shm"):
            src = GROUP_CONTAINER + suffix
            if os.path.exists(src):
                shutil.copy2(src, local + suffix)
    except PermissionError:
        print("FATAL: cannot read the calendar database.")
        print("       This is a Full Disk Access answer, not an empty calendar.")
        return None

    now = int(time.time()) - APPLE_EPOCH
    cut = now - days * 86400
    con = sqlite3.connect(f"file:{local}?mode=ro", uri=True)
    # A FUTURE EVENT IS NOT A TOUCH. Scheduling a tour is not the same as having
    # met, and counting one as contact would manufacture exactly the false
    # confidence this whole capability exists to remove. Past events become
    # inferred touches; future ones are reported separately as upcoming.
    rows = con.execute(
        """
        SELECT LOWER(p.email),
               date(ci.start_date + ?, 'unixepoch'),
               COALESCE(ci.summary, '(untitled)'),
               CASE WHEN ci.start_date <= ? THEN 'past' ELSE 'upcoming' END
        FROM CalendarItem ci
        JOIN Participant p ON p.owner_id = ci.ROWID
        WHERE p.email IS NOT NULL AND p.email <> ''
          AND ci.start_date > ?
        ORDER BY ci.start_date DESC
        """,
        (APPLE_EPOCH, now, cut),
    ).fetchall()
    con.close()
    shutil.rmtree(tmp, ignore_errors=True)
    return rows


def main():
    # --json exists so an UNATTENDED caller can act on this instead of a human
    # reading prose. bin/calendar-eventkit-capture.sh consumes it. The human
    # report is unchanged and still the default: this adds a mode, it does not
    # replace one.
    argv = [a for a in sys.argv[1:] if a != "--json"]
    as_json = "--json" in sys.argv
    days = int(argv[0]) if argv else DEFAULT_DAYS
    by_email, by_domain = load_record_contacts()
    # In --json mode stdout must be PARSEABLE and nothing else. This banner went
    # to stdout ahead of the payload and would have made json.load choke on the
    # first consumer — caught before shipping, not after.
    print(f"record contacts loaded: {len(by_email)} emails, {len(by_domain)} domains",
          file=sys.stderr if as_json else sys.stdout)

    rows = read_calendar(days)
    if rows is None:
        return 3

    latest, events, upcoming = {}, defaultdict(list), {}
    for email, day, title, when in rows:
        if when == "upcoming":
            upcoming.setdefault(email, (day, title))
            continue
        if email not in latest:
            latest[email] = day
        events[email].append((day, title))

    exact, domain, unknown, internal = {}, {}, {}, set()
    for email in latest:
        dom = email.split("@", 1)[1] if "@" in email else ""
        if dom == INTERNAL_DOMAIN:
            internal.add(email)
        elif email in by_email:
            exact[email] = by_email[email]
        elif dom in by_domain:
            domain[email] = by_domain[dom]
        else:
            unknown[email] = dom

    if as_json:
        # EXACT matches only carry a record ref, because only an exact email
        # match is evidence a named person was in the room. Domain matches say
        # "someone from that org" and must never become a dated touch on an
        # individual; they are reported for a human, never auto-logged.
        json.dump({
            "ok": True, "days": days,
            "counts": {"emails": len(latest), "internal": len(internal),
                       "exact": len(exact), "domain": len(domain),
                       "unknown": len(unknown)},
            "exact": [{"email": e, "ref": exact[e][0] if isinstance(exact[e], (list, tuple)) else exact[e],
                       "label": str(exact[e]), "last_seen": latest[e],
                       "events": [{"day": d, "title": t} for d, t in events[e][:5]]}
                      for e in exact],
            "domain": [{"email": e, "org": str(domain[e]), "last_seen": latest[e]} for e in domain],
            "unknown": [{"email": e, "domain": d, "last_seen": latest[e]}
                        for e, d in unknown.items()],
        }, sys.stdout, indent=1, default=str)
        print()
        return 0

    print(f"window: last {days} days")
    print(f"distinct attendee emails: {len(latest)}  "
          f"(internal {len(internal)}, external {len(latest) - len(internal)})")
    print()
    print(f"  EXACT match to a record contact : {len(exact)}")
    print(f"  DOMAIN match to a known org     : {len(domain)}")
    print(f"  NO match (research candidates)  : {len(unknown)}")
    print()

    if exact or domain:
        print("INFERRED TOUCHES — proposals, nothing written:")
        for tier, bucket in (("exact", exact), ("domain", domain)):
            for email, label in sorted(bucket.items(), key=lambda kv: latest[kv[0]], reverse=True):
                day, title = events[email][0]
                print(f"  [{tier:6}] {day}  {label}")
                print(f"            via {email} — {title[:60]}")
    if upcoming:
        known_up = {e: (by_email.get(e) or by_domain.get(e.split("@", 1)[1], ""))
                    for e in upcoming if not e.endswith("@" + INTERNAL_DOMAIN)}
        known_up = {e: l for e, l in known_up.items() if l}
        if known_up:
            print()
            print("UPCOMING — scheduled, NOT a touch, listed so it is never counted as one:")
            for email, label in sorted(known_up.items(), key=lambda kv: upcoming[kv[0]][0]):
                day, title = upcoming[email]
                print(f"  {day}  {label} — {title[:55]}")

    if unknown:
        print()
        print("UNMATCHED external addresses (each one a person not in the record):")
        for email, dom in sorted(unknown.items(), key=lambda kv: kv[1])[:25]:
            if dom in FREEMAIL:
                continue
            print(f"  {latest[email]}  {email}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
