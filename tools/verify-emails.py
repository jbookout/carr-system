#!/usr/bin/env python3
"""
verify-emails.py — stage 1 of the pre-send verification workflow, the free pass.

WHY THIS FILE EXISTS TWICE OVER
  `Automation/email-verification-sop.md` (2026-07-20) makes verification a standing
  rule: every outreach batch and every list added to the newsletter is verified
  BEFORE a send. It names this script as stage 1 and then says, in the SOP itself:

      "⚠️ STATUS 2026-07-20: the script is NOT in this folder."

  It was written during the Jul 19-20 bridge outage and only the note survived. So
  for a week the system carried a mandatory pre-send gate whose first stage could
  not run, and nothing said so out loud again. Rewritten 2026-07-27.

WHAT IT DOES (and deliberately does not)
  Catches the two bounce sources that are free to catch: malformed addresses, and
  domains that are dead, mistyped, or run no mail server. It resolves MX with an
  A-record fallback, because plenty of small practices accept mail on the A record.

  It NEVER SMTP-probes an individual mailbox. Probing risks the sending identity,
  and stage 2 (a paid mailbox verifier) is the step that catches a well-formed
  address nobody reads. Stage 2 needs a service account and a spend, so it is
  Joe's call and this script will not pretend to replace it.

  It writes nothing back to any record. Report in, report out.

Usage:
  run.sh verify-emails                      # every source, summary
  run.sh verify-emails --source registry    # registry | vendors | roster | csv
  run.sh verify-emails --csv path.csv --column Email
  run.sh verify-emails --segment "Web Sweep"    # substring match on any field
  run.sh verify-emails --out report.csv
"""
import argparse
import csv
import os
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from lib.drive_recovery import add_recovery_arguments, require_recovery

ap = argparse.ArgumentParser()
ap.add_argument("--source", default="all",
                choices=["all", "registry", "vendors", "roster", "csv"])
ap.add_argument("--csv")
ap.add_argument("--column", default="Email")
ap.add_argument("--segment", help="only rows containing this string in any field")
ap.add_argument("--out", help="write the per-address report to this CSV")
add_recovery_arguments(ap)
a = ap.parse_args()

try:
    vault = require_recovery(
        a, "canonical contact-email verification query API")
except ValueError as exc:
    print(f"verify-emails: STOP: {exc}", file=sys.stderr)
    raise SystemExit(2) from exc

# Deliberately strict but not RFC-pedantic: this catches the shapes that actually
# appear in hand-entered CRM data, which is where the bad rows come from.
SYNTAX = re.compile(r"^[A-Za-z0-9._%+\-']+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")

# Near-miss domains. A typo here bounces silently and looks like being ignored.
TYPO = {
    "gmail.co": "gmail.com", "gmial.com": "gmail.com", "gmai.com": "gmail.com",
    "gmail.con": "gmail.com", "gnail.com": "gmail.com", "gmail.cm": "gmail.com",
    "yahoo.co": "yahoo.com", "yaho.com": "yahoo.com", "yahooo.com": "yahoo.com",
    "hotmail.co": "hotmail.com", "hotmial.com": "hotmail.com",
    "outlook.co": "outlook.com", "outook.com": "outlook.com",
    "aol.co": "aol.com", "iclould.com": "icloud.com", "icloud.co": "icloud.com",
    "comcast.com": "comcast.net", "bellsouth.com": "bellsouth.net",
}

_dns_cache: dict[str, tuple[bool, str]] = {}


def has_mail(domain):
    """(ok, how). MX first, A as fallback — small practices often have only an A."""
    d = domain.lower().strip(".")
    if d in _dns_cache:
        return _dns_cache[d]
    def dig(rtype):
        try:
            out = subprocess.run(["dig", "+short", "+time=3", "+tries=1", rtype, d],
                                 capture_output=True, text=True, timeout=8)
            return [l for l in out.stdout.strip().splitlines() if l.strip()]
        except Exception:
            return []
    res = (True, "MX") if dig("MX") else \
          (True, "A (no MX)") if dig("A") else (False, "no MX and no A")
    _dns_cache[d] = res
    return res


def sheet_rows(path, sheet):
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip() for h in next(it, [])]
    out = [dict(zip(hdr, r)) for r in it if any(x is not None for x in r)]
    wb.close()
    return out


SOURCES = {
    "registry": ("DNA/Leads/lead-registry.xlsx", "Registry", "Lead ID", "Contact Name", "Email"),
    "vendors":  ("DNA/Network/vendors.xlsx", "Vendors", "ID", "Name", "Email"),
    "roster":   ("DNA/Clients/client-roster.xlsx", "Clients", "Client ID", "Name", "Email"),
}

records = []
if a.source == "csv" or a.csv:
    if not a.csv:
        raise SystemExit("verify-emails: --source csv needs --csv PATH.")
    with open(a.csv, newline="", encoding="utf-8-sig") as fh:
        for i, row in enumerate(csv.DictReader(fh), 2):
            records.append(("csv", str(i), row.get("Name", ""), row.get(a.column, ""), row))
else:
    for key, (rel, sheet, idc, namec, mailc) in SOURCES.items():
        if a.source not in ("all", key):
            continue
        for r in sheet_rows(os.path.join(vault, rel), sheet):
            records.append((key, str(r.get(idc) or ""), str(r.get(namec) or ""),
                            str(r.get(mailc) or ""), r))

if a.segment:
    seg = a.segment.lower()
    records = [t for t in records
               if any(seg in str(v).lower() for v in t[4].values() if v is not None)]

# Real cells are messy. Many carry the address PLUS a hand-typed note:
#   "j.sample@example.net *(added Jul 6, 2026 — Dell's export)*"
#   "mayyybe a.person_cpa@example.com"
#   "person1@example.org \ person2@example.org (or possibly just person3@example.org)"
# (examples sanitized 2026-08-06, ORDER 42b — the originals were real lead emails)
# Splitting on punctuation shreds those and reports the fragments as malformed,
# which is a checker inventing its own findings. Extract addresses instead, and
# report a cell that holds text but no address as its own (real) category.
FIND = re.compile(r"[A-Za-z0-9._%+\-']+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

report, counts = [], {"ok": 0, "syntax": 0, "typo": 0, "dead": 0, "blank": 0,
                      "multi": 0, "noaddr": 0}
for src, ident, name, raw, _row in records:
    raw = (raw or "").strip()
    if not raw:
        counts["blank"] += 1
        continue
    addrs = []
    for m in FIND.finditer(raw):
        addr = m.group(0).rstrip(".-_'")
        if addr not in addrs:
            addrs.append(addr)
    if not addrs:
        counts["noaddr"] += 1
        report.append((src, ident, name, raw[:60], "NO ADDRESS",
                       "cell holds text but no address"))
        continue
    if len(addrs) > 1:
        counts["multi"] += 1
    for addr in addrs:
        if not SYNTAX.match(addr):
            counts["syntax"] += 1
            report.append((src, ident, name, addr, "MALFORMED", "not a valid address"))
            continue
        dom = addr.split("@")[1].lower()
        if dom in TYPO:
            counts["typo"] += 1
            report.append((src, ident, name, addr, "TYPO DOMAIN", "did you mean %s" % TYPO[dom]))
            continue
        ok, how = has_mail(dom)
        if ok:
            counts["ok"] += 1
            report.append((src, ident, name, addr, "VALID", how))
        else:
            counts["dead"] += 1
            report.append((src, ident, name, addr, "DEAD DOMAIN", how))

print("verify-emails — stage 1, the free domain pass")
print("  source: %s%s" % (a.source, " · segment %r" % a.segment if a.segment else ""))
print("  %d record(s), %d with an address, %d distinct domain(s)"
      % (len(records), len(records) - counts["blank"], len(_dns_cache)))
print()
print("  VALID .......... %d" % counts["ok"])
print("  MALFORMED ...... %d" % counts["syntax"])
print("  TYPO DOMAIN .... %d" % counts["typo"])
print("  DEAD DOMAIN .... %d" % counts["dead"])
print("  no address ..... %d  (cell empty)" % counts["blank"])
print("  NO ADDRESS ..... %d  (cell has text but no address in it)" % counts["noaddr"])
if counts["multi"]:
    print("  (%d row(s) held more than one address; each was checked separately)" % counts["multi"])

bad = [r for r in report if r[4] != "VALID"]
if bad:
    print()
    print("  Problems:")
    for src, ident, name, addr, verdict, why in bad:
        print("    %-8s %-8s %-26s %-34s %s — %s"
              % (src, ident, name[:26], addr[:34], verdict, why))

if a.out:
    with open(a.out, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Source", "ID", "Name", "Email", "Verdict", "Detail"])
        w.writerows(report)
    print("\n  wrote %s (%d rows)" % (a.out, len(report)))

print()
print("  Stage 1 only. A VALID verdict means the domain accepts mail, NOT that this")
print("  mailbox exists or is read. The mailbox pass (stage 2, paid) is still the gate")
print("  immediately before any send — see Automation/email-verification-sop.md.")
sys.exit(0)
