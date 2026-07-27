#!/usr/bin/env python3
"""
registry-audit.py — read-only integrity check on lead-registry.xlsx.

Answers the question nobody had a way to ask before 2026-07-27: does every
pointer into the registry still resolve to the person it was written for?

Checks, in order of how much damage they represent:

  1  POINTER ROT — every "Registry: L-xxx" in a prospect dossier and every
     L-ID cited in clients-active.md is resolved against the actual row. A
     pointer to a missing ID, or to a row occupied by a different practice, is
     reported with both sides shown. This is the check that would have caught
     C-118 … C-121 pointing at rows that were never theirs.

  2  OCCUPANT DRIFT — the newest backup is diffed against the live file. Any
     existing Lead ID whose Contact Name or Practice changed is flagged. This is
     the #101 signature after the fact.

  3  SCHEMA — all 26 canonical columns present. (#100.)

  4  LEDGER STALENESS — the highest Lead ID recorded in the Intake Log against
     the highest actually in the Registry sheet. A ledger behind the sheet is
     what let the 7/25 writer allocate on top of live rows: it took max+1 from
     the ledger because the 7/14 and 7/16 adds were never logged.

  5  DUPES AND GAPS — duplicate IDs are errors. Gaps are reported as tombstones,
     with retired IDs named so they are not re-litigated every audit.

Writes nothing. Exit 1 if any ERROR fired, 0 otherwise (WARN does not fail).

Usage:  run.sh registry-audit [--verbose]
"""
import os
import re
import sys
import glob

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "lib"))
import openpyxl
from registry import (REGISTRY_COLUMNS, RETIRED_IDS, id_num, fmt_id, _s)

VERBOSE = "--verbose" in sys.argv
ROOT = os.environ.get("CARR_VAULT") or (sys.argv[1] if len(sys.argv) > 1 and not sys.argv[1].startswith("-") else None)
if not ROOT:
    raise SystemExit("registry-audit: pass the vault root or set CARR_VAULT.")

REG = os.path.join(ROOT, "DNA", "Leads", "lead-registry.xlsx")
PROSPECTS = os.path.join(ROOT, "DNA", "Clients", "prospects")
ROSTER = os.path.join(ROOT, "DNA", "Clients", "clients-active.md")

errors, warns = [], []


def err(msg):
    errors.append(msg)
    print("  ERROR  " + msg)


def warn(msg):
    warns.append(msg)
    print("  WARN   " + msg)


def ok(msg):
    if VERBOSE:
        print("  ok     " + msg)


STOP = {"the", "and", "for", "llc", "pllc", "inc", "pa", "pc", "dds", "dmd", "md",
        "dr", "do", "lpc", "arnp", "ms", "facs", "center", "centre", "health",
        "medical", "dental", "clinic", "practice", "group", "services", "associates",
        "care", "family", "of", "at", "in", "new", "north", "south", "east", "west"}


def tokens(*parts):
    """Word set for loose identity comparison.

    Dossier filenames are CamelCase-and-hyphen ("GulfCoastPelvicFloor",
    "PremierHealthWellness-RandallMacDonnell"), so a naive lowercase split makes
    one giant token that matches nothing and reports a false mismatch. Split on
    the case boundary first.
    """
    blob = " ".join(_s(p) for p in parts)
    blob = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", blob)
    blob = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", " ", blob)
    return {w for w in re.findall(r"[a-z]{4,}", blob.lower()) if w not in STOP}


# ---------------------------------------------------------------- load

if not os.path.exists(REG):
    raise SystemExit("registry-audit: no registry at %s" % REG)

wb = openpyxl.load_workbook(REG, data_only=True)
ws = wb["Registry"]
hdr = [_s(c.value) for c in ws[1]]
colmap = {h: i for i, h in enumerate(hdr) if h}

rows = {}
dupes = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not any(x is not None and _s(x) for x in r):
        continue
    n = id_num(r[colmap["Lead ID"]]) if "Lead ID" in colmap else None
    if n is None:
        continue
    if n in rows:
        dupes.append(n)
    rows[n] = {h: r[i] for h, i in colmap.items() if i < len(r)}

print("registry-audit — %d rows, %d columns" % (len(rows), len(hdr)))
print()

# ---------------------------------------------------------------- 3 schema

print("[3] schema")
missing = [c for c in REGISTRY_COLUMNS if c not in colmap]
if missing:
    err("canonical column(s) missing: %s (defect #100 class)" % missing)
else:
    ok("all %d canonical columns present" % len(REGISTRY_COLUMNS))
print()

# ---------------------------------------------------------------- 5 dupes/gaps

print("[5] ids")
if dupes:
    err("duplicate Lead ID(s): %s" % [fmt_id(n) for n in sorted(set(dupes))])
else:
    ok("no duplicate ids")
if rows:
    lo, hi = min(rows), max(rows)
    gaps = [n for n in range(lo, hi + 1) if n not in rows]
    for n in gaps:
        if n in RETIRED_IDS:
            ok("%s gap — retired by design, not a loss" % fmt_id(n))
        else:
            warn("%s is a gap and is NOT on the retired list — was a row lost, "
                 "or should it be retired explicitly?" % fmt_id(n))
print()

# ---------------------------------------------------------------- 4 ledger

print("[4] ledger vs sheet")
led_max = 0
if "Intake Log" in wb.sheetnames:
    for r in wb["Intake Log"].iter_rows(values_only=True):
        for cell in r:
            for m in re.finditer(r"L-0*(\d+)", _s(cell)):
                led_max = max(led_max, int(m.group(1)))
if rows:
    sheet_max = max(rows)
    if led_max < sheet_max:
        warn("Intake Log's highest recorded id is %s but the sheet reaches %s. "
             "%d add(s) were never ledgered. A writer that allocates off the ledger "
             "will land on live rows — this is exactly what happened on 2026-07-25."
             % (fmt_id(led_max), fmt_id(sheet_max), sheet_max - led_max))
    else:
        ok("ledger current (ledger %s, sheet %s)" % (fmt_id(led_max), fmt_id(sheet_max)))
print()

# ---------------------------------------------------------------- 2 drift

print("[2] occupant drift vs newest backup")
backups = sorted(glob.glob(os.path.join(os.path.dirname(REG), "lead-registry.backup-*.xlsx")),
                 key=os.path.getmtime)
if not backups:
    warn("no backup to diff against")
else:
    bpath = backups[-1]
    bwb = openpyxl.load_workbook(bpath, data_only=True)
    bws = bwb["Registry"]
    bhdr = [_s(c.value) for c in bws[1]]
    bcol = {h: i for i, h in enumerate(bhdr) if h}
    drift = 0
    if "Lead ID" in bcol:
        for r in bws.iter_rows(min_row=2, values_only=True):
            if not any(x is not None and _s(x) for x in r):
                continue
            n = id_num(r[bcol["Lead ID"]])
            if n is None:
                continue
            was = (_s(r[bcol["Contact Name"]]), _s(r[bcol["Practice"]]))
            if n not in rows:
                err("%s present in %s, absent now — row lost"
                    % (fmt_id(n), os.path.basename(bpath)))
                drift += 1
                continue
            now = (_s(rows[n].get("Contact Name")), _s(rows[n].get("Practice")))
            if was != now and not tokens(*was) & tokens(*now):
                err("%s changed occupant: was %s / %s, now %s / %s"
                    % (fmt_id(n), was[0] or "(blank)", was[1] or "(blank)",
                       now[0] or "(blank)", now[1] or "(blank)"))
                drift += 1
    bwb.close()
    if not drift:
        ok("no drift vs %s" % os.path.basename(bpath))
print()

# ---------------------------------------------------------------- 1 pointers

print("[1] pointer rot")
pointers = []

for path in sorted(glob.glob(os.path.join(PROSPECTS, "*.md"))):
    txt = open(path, encoding="utf-8", errors="replace").read()
    head = txt[:4000]
    found = set()
    for m in re.finditer(r"[Rr]egistry:?\**\s*:?\s*\**\s*(L-\d{3})", head):
        found.add(m.group(1))
    for m in re.finditer(r"^#\s+.*?\((L-\d{3})\)", head, re.M):
        found.add(m.group(1))
    for lid in found:
        pointers.append((os.path.basename(path), lid, os.path.splitext(os.path.basename(path))[0]))

if os.path.exists(ROSTER):
    for line in open(ROSTER, encoding="utf-8", errors="replace"):
        if not line.startswith("|"):
            continue
        for m in re.finditer(r"(L-\d{3})", line):
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            label = cells[1] if len(cells) > 1 else line[:60]
            pointers.append(("clients-active.md", m.group(1), label))

if not pointers:
    warn("no registry pointers found to check")

seen = set()
for src, lid, label in pointers:
    if (src, lid) in seen:
        continue
    seen.add((src, lid))
    n = id_num(lid)
    if n in RETIRED_IDS:
        err("%s -> %s, which is a RETIRED id. That pointer can never resolve. "
            "(%s)" % (src, lid, label))
        continue
    if n not in rows:
        err("%s -> %s, which has no row in the registry. (%s)" % (src, lid, label))
        continue
    row_tok = tokens(rows[n].get("Contact Name"), rows[n].get("Practice"))
    ptr_tok = tokens(label.replace("-", " "))
    if ptr_tok and row_tok and not (ptr_tok & row_tok):
        err("%s -> %s, but %s holds %s / %s. The pointer names %s. Different people."
            % (src, lid, lid, _s(rows[n].get("Contact Name")) or "(blank)",
               _s(rows[n].get("Practice")) or "(blank)", label))
    else:
        ok("%s -> %s resolves" % (src, lid))

print()

# ---------------------------------------------------------------- 6 detail files

print("[6] detail file paths")
# The July 11 refactor moved every dossier from top-level Prospects/ to
# DNA/Clients/prospects/. Eight rows (L-001 … L-008, the oldest and most-worked
# leads in the system) still carried the old path 16 days later, and nothing said
# so. A pointer that names a file is worth exactly as much as the file existing.
checked = broken = prose = 0
for n in sorted(rows):
    df = _s(rows[n].get("Detail File"))
    if not df:
        continue
    # Some rows use this column for a narrative note rather than a path. That is
    # pre-existing convention, not a defect; count it and move on.
    if "/" not in df or len(df) > 160 or " " in df.split("/")[0]:
        prose += 1
        continue
    checked += 1
    if not os.path.exists(os.path.join(ROOT, df)):
        broken += 1
        err("%s Detail File does not exist: %s" % (fmt_id(n), df))
if not broken:
    ok("%d detail path(s) resolve (%d rows hold prose, not a path)" % (checked, prose))
print()
print("registry-audit: %d error(s), %d warning(s)" % (len(errors), len(warns)))
wb.close()
sys.exit(1 if errors else 0)
