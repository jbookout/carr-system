#!/usr/bin/env python3
"""
graph-health.py — data-quality anomalies as a LIST, not as dots to click.

WHY
  The Obsidian graph is good at making you ASK "why isn't that connected?" and bad
  at answering it. Every diagnosis in the 2026-07-25 session came from the command
  line, not from the graph. This gives the graph its job back (orientation, shape)
  by reporting the anomalies it can only gesture at.

  Reads the same four sources as build-graph-notes.py plus the generated Graph/,
  so it stays honest about what the graph is actually showing.

WHAT IT CHECKS
  1  Isolated nodes .......... records with no edge at all — usually nothing real in them
  2  Placeholder names ....... TBD / unknown / "last name" / parentheticals
  3  Multi-person fields ..... two people crammed into one name cell
  4  Cross-record duplicates . same person as BOTH a lead and a client — the failure
                               that put paying clients on a prospecting drip
  5  Missing source .......... leads with no source, deals with no referral
  6  Name vs email ........... email local part shares no token with the name —
                               catches the "emailing the wrong human" class
  7  Duplicate node titles ... regression guard; must stay 0 or links go ambiguous

Usage:  run.sh graph-health   |   python3 pipelines/graph-health.py [VAULT] [--verbose]
Exit 0 always — this is a report, not a gate.
"""
import sys, os, re, json, glob
from collections import defaultdict
import openpyxl

args = [a for a in sys.argv[1:] if not a.startswith("--")]
VERBOSE = "--verbose" in sys.argv
ROOT = args[0] if args else os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", ".."))
GRAPH = os.path.join(ROOT, "Graph")

def s(v): return str(v if v is not None else "").strip()

def rows(path, sheet):
    if not os.path.exists(path): return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames: wb.close(); return []
    ws = wb[sheet]; it = ws.iter_rows(values_only=True)
    hdr = [s(h) for h in next(it, [])]
    out = [dict(zip(hdr, r)) for r in it if any(x is not None for x in r)]
    wb.close(); return out

vendors = rows(os.path.join(ROOT, "DNA/Network/vendors.xlsx"), "Vendors")
leads   = rows(os.path.join(ROOT, "DNA/Leads/lead-registry.xlsx"), "Registry")
clients = rows(os.path.join(ROOT, "DNA/Clients/client-roster.xlsx"), "Clients")
deals   = []
dp = os.path.join(ROOT, "DNA/Deal Management/panhandle-team-deals.json")
if os.path.exists(dp): deals = json.load(open(dp)).get("deals", [])

findings = []          # (severity, category, detail)
def add(sev, cat, detail): findings.append((sev, cat, detail))

# ---------- 1 & 7: graph-derived checks ----------
titles = defaultdict(list)
for sub in ("vendors", "leads", "clients", "deals"):
    for f in glob.glob(os.path.join(GRAPH, sub, "*.md")):
        titles[os.path.basename(f)].append(sub)

for name, subs in sorted(titles.items()):
    if len(subs) > 1:
        add("HIGH", "Duplicate node title",
            f"{name[:-3]} exists in {' + '.join(subs)} — every [[link]] to it is ambiguous")

ent = {}
for sub in ("vendors", "leads", "clients", "deals"):
    for f in glob.glob(os.path.join(GRAPH, sub, "*.md")):
        ent[os.path.splitext(os.path.basename(f))[0]] = 0
for sub in ("vendors", "leads", "clients", "deals", "hubs"):
    for f in glob.glob(os.path.join(GRAPH, sub, "*.md")):
        t = open(f, encoding="utf-8", errors="ignore").read()
        src = os.path.splitext(os.path.basename(f))[0]
        for m in re.findall(r"\[\[([^\]]+)\]\]", t):
            if m in ent:
                ent[m] += 1
                if src in ent: ent[src] += 1
for n, c in sorted(ent.items()):
    if c == 0:
        add("MED", "Isolated node", f"{n} — no firm, market, owner or referral to connect on")

# ---------- 2 & 3: name quality ----------
PLACEHOLDER = re.compile(r"\bTBD\b|\bunknown\b|last name|\benrich\b|^\(|linkedin|listing \d+", re.I)
SUFFIX = r"(?:jr|sr|ii|iii|dds|dmd|md|do|pa|np|cpa|esq|phd|rn|fnp|otr)"
MULTINAME = re.compile(r"^([A-Z][a-z'’-]+ [A-Z][a-z'’-]+) ([A-Z][a-z'’-]+ [A-Z][a-z'’-]+)$")

# A practice name is four capitalised words too ("Bay Area Oral Surgery"), so the
# four-word shape alone is not evidence of two people. Any business/clinical word
# means it is an entity, not a pair of humans.
ORG_WORD = re.compile(
    r"\b(center|centre|clinic|clinical|surgery|surgical|health|healthcare|medical|"
    r"medicine|dental|dentistry|dentist|ortho\w*|chiro\w*|veterinar\w*|animal|pet|"
    r"vision|eye|optical|therapy|therapies|wellness|rehab\w*|hospital|institute|"
    r"practice|group|associates|partners|family|pediatric\w*|women\w*|floor|spa|"
    r"studio|lab|labs|imaging|urgent|care|physicians?|dermatolog\w*|endodont\w*|"
    r"periodont\w*|oral|maxillofacial|podiatr\w*|cardiolog\w*|ent|dpc|pllc|llc|inc|"
    r"pa|pc|bank|financial|insurance|realty|construction|design|build|solutions)\b", re.I)

def name_checks(label, name, ident):
    if not name: return
    if PLACEHOLDER.search(name):
        add("MED", "Placeholder name", f"{label} {ident}: “{name}”")
    m = MULTINAME.match(name.strip())
    if m and not ORG_WORD.search(name) and not re.fullmatch(SUFFIX, m.group(2).split()[-1], re.I):
        add("HIGH", "Two people in one field",
            f"{label} {ident}: “{name}” looks like “{m.group(1)}” + “{m.group(2)}”")

for v in vendors: name_checks("vendor", s(v.get("Name")), s(v.get("ID")))
for l in leads:   name_checks("lead",   s(l.get("Contact Name")), s(l.get("Lead ID")))
for c in clients: name_checks("client", s(c.get("Name")), s(c.get("Client ID")))

# ---------- 4: same person as both lead and client ----------
lead_names  = {s(l.get("Contact Name")).lower(): s(l.get("Lead ID")) for l in leads if s(l.get("Contact Name"))}
for c in clients:
    nm = s(c.get("Name")).lower()
    if nm and nm in lead_names:
        add("HIGH", "Lead + client duplicate",
            f"“{s(c.get('Name'))}” is {s(c.get('Client ID'))} (client, {s(c.get('Status'))}) "
            f"AND {lead_names[nm]} (lead) — check the lead is not on a drip")

# ---------- 5: missing source ----------
no_src = [s(l.get("Lead ID")) for l in leads
          if not s(l.get("Source Detail (V-ID / event / referrer)")) and not s(l.get("Source Type"))]
if no_src:
    add("LOW", "Lead with no source", f"{len(no_src)} leads: {', '.join(no_src[:12])}"
        + (" …" if len(no_src) > 12 else ""))
no_ref = [s(d.get("name")) for d in deals if not s(d.get("referral"))]
if no_ref:
    add("LOW", "Deal with no referral", f"{len(no_ref)} deals: {', '.join(no_ref[:10])}"
        + (" …" if len(no_ref) > 10 else ""))

# ---------- 6: name vs email ----------
TITLES = {"dr", "mr", "mrs", "ms", "dds", "dmd", "md", "do", "jr", "sr", "ii", "iii",
          "cpa", "esq", "phd", "rn", "np", "pa", "fnp", "otr", "the", "and"}

def name_parts(name):
    return [p for p in re.split(r"[^a-zA-Z]+", name.lower())
            if len(p) > 1 and p not in TITLES]

def email_check(label, name, email, ident, company=""):
    """Flag only when the address plausibly belongs to SOMEONE ELSE.

    Real addresses take many shapes and none of them are a token match:
      nileshpatel@   (concatenated)   jholder@  (initial+surname)
      hicksc@        (surname+initial) c.busby@ (initial.surname)
    Flagging those produced 207 false positives on the first run. Only report when
    no part of the person's name appears in the local part in ANY of those forms —
    which is what an address like rachel.noell@ on a row named Tatum Cannon looks
    like, and that one is a real defect worth catching.
    """
    name, email = s(name), s(email)
    if not name or "@" not in email or PLACEHOLDER.search(name): return
    parts = name_parts(name)
    if not parts: return
    local = re.sub(r"[^a-z]", "", email.split("@")[0].lower())
    if not local: return
    for p in parts:
        if p in local or local in p:                    # concatenated or truncated
            return
    initials = {p[0] for p in parts}
    for p in parts:                                     # initial+surname / surname+initial
        if len(p) > 2 and any(local == i + p or local == p + i or
                              local.startswith(i + p) or local.startswith(p + i)
                              for i in initials):
            return
    # A shared inbox (southwoodconst@, terraequities@, monarchbc@) is not a defect —
    # it is how small firms operate. If the local part echoes the company or the
    # domain, treat it as a generic address and note it separately.
    org = re.sub(r"[^a-z]", "", (company or "").lower())
    dom = re.sub(r"[^a-z]", "", email.split("@")[1].split(".")[0].lower())
    generic = (org and (local in org or org.startswith(local) or local.startswith(org[:6]))) \
              or (dom and (local in dom or dom.startswith(local)))
    if generic:
        add("LOW", "Shared/company inbox, not a personal address",
            f"{label} {ident}: “{name}” → {email}")
        return
    add("HIGH", "Name vs email disagree",
        f"{label} {ident}: “{name}” but address is {email} — mail may reach a different person")

for v in vendors: email_check("vendor", v.get("Name"), v.get("Email"), s(v.get("ID")), s(v.get("Company")))
for l in leads:   email_check("lead",   l.get("Contact Name"), l.get("Email"), s(l.get("Lead ID")), s(l.get("Practice")))
for c in clients: email_check("client", c.get("Name"), c.get("Email"), s(c.get("Client ID")), s(c.get("Practice / Entity")))

# ---------- report ----------
ORDER = {"HIGH": 0, "MED": 1, "LOW": 2}
by_cat = defaultdict(list)
for sev, cat, detail in findings: by_cat[(ORDER[sev], sev, cat)].append(detail)

print(f"\nGRAPH HEALTH — {len(vendors)} vendors, {len(leads)} leads, "
      f"{len(clients)} clients, {len(deals)} deals\n" + "=" * 72)
if not findings:
    print("\nNo anomalies found.\n"); sys.exit(0)

CAP = 10000 if VERBOSE else 6
for key in sorted(by_cat):
    _, sev, cat = key
    items = by_cat[key]
    print(f"\n[{sev}] {cat} — {len(items)}")
    for d in items[:CAP]:
        print(f"   · {d}")
    if len(items) > CAP:
        print(f"   … {len(items)-CAP} more (--verbose for all)")

hi = sum(1 for f in findings if f[0] == "HIGH")
print("\n" + "=" * 72)
print(f"{len(findings)} findings — {hi} HIGH."
      + ("" if VERBOSE else "  Re-run with --verbose for the full list."))
print("HIGH = fix before the data is trusted for outreach or a merge.\n")
