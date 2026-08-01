#!/usr/bin/env python3
"""
parse_party_links.py — ORDER 17 / amendment 4: turn `vendor.links_label` prose
into real `party_link` edges, so who-do-we-know and the reciprocity ledger stop
being questions somebody has to answer from memory.

WHY THIS IS A PIPELINE AND NOT A VERB (Fable ruling (a), 2026-07-31)
  This is BACKFILL of frozen imported text, the same class of work as
  `import_wave1.py`: one pass, direct DB as writer, a dry-run report reviewed
  BEFORE anything is written. Links created from here on go through the
  `link-parties` verb like any other record change. This script is not part of
  daily traffic and should not be rerun as a habit.

THE GRAMMAR, VERBATIM FROM THE ORDER (ruling (b)) — nothing beyond it
  * A token matching  V-[A-Z]{3}-\\d+ | T-\\d+ | C-\\d+ | L-\\d+  resolves
    MECHANICALLY to its party via the ref (T-### is a vendor ref too — the
    'Target (not yet met)' tier, 41 of them live).
  * A bare personal/org name resolves ONLY on exact unique match against
    `party.name`.
  * Anything else — no match, multiple matches, prose that isn't a link — is
    AMBIGUOUS, and ambiguous is NEVER guessed (ruling (c)).

WHAT AMBIGUOUS COSTS (ruling (c))
  One `event` row per vendor carrying the unparsed remainder, plus the full list
  in the dry-run report for a human review pass. The nine-orphans pattern from
  amendment 2 is the precedent: the record self-flags, a human decides.

EDGE SHAPE (ruling (d))
  party_link(from_party = the vendor's party, to_party = the resolved party,
  kind = 'intro' | 'referral' | 'knows', note = the original label VERBATIM).
  Default kind is 'knows' where the label names a party without a relation verb.
  Idempotent on (from_party, to_party, kind): a rerun writes zero rows.

WHAT THIS NEVER DOES (ruling (e) + the order's stop rules)
  `links_label` is NOT cleared, rewritten, or deleted. It is a shim with a death
  sentence per amendment 5 and it dies at the vendors-file repoint, not here.

ROW ACCOUNTING (ORDER 32(a), 2026-07-31)
  ORDER 17 counted ITEMS — 28 edges, 4 ambiguous — which is the right number for
  "did the grammar work" and the wrong number for "is every Links row accounted
  for". One row can produce three edges and one review item at once (V-CPA-022
  does exactly that), so item counts cannot tell you whether a row fell through
  the floor. Every vendor row now lands in exactly one of three buckets — EDGED,
  REVIEW-ONLY, EMPTY — and the three are asserted to sum to the vendor count. An
  accounting that does not sum is not an accounting, so a mismatch is a non-zero
  exit, not a footnote.

THE XLSX CROSS-CHECK (ORDER 32(a))
  `DNA/Network/vendors.xlsx` is a GENERATED file whose `Links` column renders
  `vendor.links_label`, so the two should be the same rows. Should is not is: a
  stale export, or a hand edit somebody made to the sheet, would put Links text
  in front of a human that the parser never sees. `--xlsx <path>` reads the sheet
  READ-ONLY and reports drift in both directions. It is a check, never a second
  input — nothing is ever parsed out of the file.

ALREADY-RESOLVED REVIEW ITEMS
  A review item is not open for ever. Three of ORDER 17's four ("Rick
  McClanahan", who is Ric McClanahan, V-BNK-030) were settled by Joe and written
  with `link-parties`, so the edges exist and the grammar still cannot see them —
  it re-flags the same prose every run. The report now marks an ambiguous item
  RESOLVED when a verb-written edge already leaves that vendor, so Joe's review
  list is the work that is actually left rather than the work that was once left.

Usage:
  DATABASE_URL=... .venv/bin/python pipelines/parse_party_links.py           # dry run (default)
  DATABASE_URL=... .venv/bin/python pipelines/parse_party_links.py --apply   # write
  ... --xlsx "/path/to/vendors.xlsx"                                         # + drift check
Writes out/party-link-parse-<stamp>.md either way. The dry run touches nothing.
"""

import argparse
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import psycopg

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "out"

# ── the grammar, ruling (b) ──────────────────────────────────────────────────
REF_RE = re.compile(r"\b(?:V-[A-Z]{3}-\d+|T-\d+|C-\d+|L-\d+)\b")
# Items inside one label. ' / ' only with surrounding whitespace, so a slash
# inside a name ("Tubbs / George DVM" is a real party name) cannot split it by
# accident when it is written without spaces.
ITEM_SPLIT_RE = re.compile(r"\s*(?:;|·|\||\n|(?<=\s)/(?=\s))\s*")
PAREN_RE = re.compile(r"\([^)]*\)")
# Prose after a dash separator: 'Chris Kelly (V-CPA-006) — offered intro'
DASH_CUT_RE = re.compile(r"\s+[—–-]\s+.*$")

PROVENANCE_SOURCE = "links_label_parse"   # party_link.source is NOT NULL and is
                                          # a provenance column; this says exactly
                                          # where the row came from and collides
                                          # with neither 'stated' (verb-written)
                                          # nor 'import' (wave-1 records).
REVIEW_VERB = "amendment-4-review"


def kind_for(text):
    """Ruling (d): relation verb decides; 'knows' is the default."""
    t = text.lower()
    if "intro" in t:
        return "intro"
    if "refer" in t:
        return "referral"
    return "knows"


def name_candidate(item):
    """The bare-name form of an item: parentheticals and post-dash prose removed.

    Deliberately conservative — it strips decoration, it never rewrites a name.
    """
    s = PAREN_RE.sub(" ", item)
    s = DASH_CUT_RE.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.strip(" .,:")
    return s


def load(conn):
    """Everything the parse needs, read once."""
    cur = conn.execute("""
        select v.id, v.vendor_ref, p.id, p.name, coalesce(org.name,''), v.links_label
          from vendor v
          join party p on p.id = v.party_id
          left join party org on org.id = p.org_id
         where coalesce(trim(v.links_label),'') <> ''
         order by v.vendor_ref
    """)
    vendors = cur.fetchall()

    # [ORDER 32] The denominator. Row accounting has to sum to EVERY vendor row,
    # not just the ones that happen to carry text, or "explicitly empty" is a
    # claim nobody checked.
    all_vendors = conn.execute("""
        select v.vendor_ref, coalesce(trim(v.links_label),'') <> '' as has_links
          from vendor v order by v.vendor_ref
    """).fetchall()

    refs = {}          # ref -> (party_id, display, kind_of_record)
    for r, pid in conn.execute(
            "select vendor_ref, party_id from vendor where vendor_ref is not null").fetchall():
        refs[r] = (pid, "vendor")
    for r, pid in conn.execute(
            "select roster_ref, party_id from client where roster_ref is not null").fetchall():
        refs.setdefault(r, (pid, "client"))
    for r, pid in conn.execute(
            "select registry_ref, party_id from lead where registry_ref is not null").fetchall():
        refs.setdefault(r, (pid, "lead"))

    names = {}         # lower(name) -> [party_id, ...]
    for pid, nm in conn.execute("select id, name from party").fetchall():
        names.setdefault(nm.strip().lower(), []).append(pid)

    party_name = {pid: nm for pid, nm in conn.execute("select id, name from party").fetchall()}

    existing = set()
    for f, t, k in conn.execute("select from_party, to_party, kind from party_link").fetchall():
        existing.add((f, t, k))

    reviewed = {r[0] for r in conn.execute(
        "select subject_id from event where verb = %s and subject_type = 'vendor'",
        (REVIEW_VERB,)).fetchall()}

    # [ORDER 32] Parties that have gained a VERB-written edge since the backfill.
    # `source='stated'` is link-parties' provenance, so an outbound stated edge is
    # Joe having answered a question the grammar could not. Used only to mark a
    # review item resolved in the report — never to create or suppress an edge.
    stated_from = {r[0] for r in conn.execute(
        "select distinct from_party from party_link where source = 'stated'").fetchall()}

    return vendors, all_vendors, refs, names, party_name, existing, reviewed, stated_from


def parse(vendors, refs, names, party_name):
    """Returns (edges, ambiguous). Pure — no DB, no writes."""
    edges = []          # dicts: from_party, to_party, kind, note, how, vendor_ref, item, target
    ambiguous = []      # dicts: vendor_ref, vendor_name, item, reason, label
    seen = set()        # in-run dedup on (from, to, kind)

    for vid, vref, vparty, vname, vorg, label in vendors:
        for raw_item in ITEM_SPLIT_RE.split(label):
            item = raw_item.strip()
            if not item:
                continue
            k = kind_for(item)
            found = REF_RE.findall(item)

            if found:
                for ref in found:
                    hit = refs.get(ref)
                    if not hit:
                        ambiguous.append(dict(
                            vendor_ref=vref, vendor_name=vname, item=item, label=label,
                            reason=f"ref {ref} resolves to no record"))
                        continue
                    tparty, rec_kind = hit
                    if tparty == vparty:
                        ambiguous.append(dict(
                            vendor_ref=vref, vendor_name=vname, item=item, label=label,
                            reason=f"ref {ref} resolves to this vendor's own party (self-link)"))
                        continue
                    key = (vparty, tparty, k)
                    if key in seen:
                        continue
                    seen.add(key)
                    edges.append(dict(
                        from_party=vparty, to_party=tparty, kind=k, note=label,
                        how="mechanical", vendor_ref=vref, vendor_name=vname, item=item,
                        target=f"{ref} {party_name.get(tparty,'?')}"))
                continue

            cand = name_candidate(item)
            if not cand:
                ambiguous.append(dict(
                    vendor_ref=vref, vendor_name=vname, item=item, label=label,
                    reason="no ref token and nothing name-shaped left after decoration"))
                continue
            hits = names.get(cand.lower(), [])
            if len(hits) == 1:
                tparty = hits[0]
                if tparty == vparty:
                    ambiguous.append(dict(
                        vendor_ref=vref, vendor_name=vname, item=item, label=label,
                        reason=f"'{cand}' is this vendor's own party (self-link)"))
                    continue
                key = (vparty, tparty, k)
                if key in seen:
                    continue
                seen.add(key)
                edges.append(dict(
                    from_party=vparty, to_party=tparty, kind=k, note=label,
                    how="name-matched", vendor_ref=vref, vendor_name=vname, item=item,
                    target=party_name.get(tparty, "?")))
            elif len(hits) == 0:
                ambiguous.append(dict(
                    vendor_ref=vref, vendor_name=vname, item=item, label=label,
                    reason=f"'{cand}' matches no party name exactly"))
            else:
                ambiguous.append(dict(
                    vendor_ref=vref, vendor_name=vname, item=item, label=label,
                    reason=f"'{cand}' matches {len(hits)} parties — ambiguous by definition"))

    return edges, ambiguous


def row_accounting(all_vendors, vendors, edges, ambiguous):
    """[ORDER 32] Place EVERY vendor row in exactly one bucket.

    edged        — the row produced at least one edge (it may ALSO have produced a
                   review item; an edge is the stronger fact, so the row counts as
                   handled and the leftover item still shows in the review list)
    review_only  — the row produced no edge at all and at least one review item
    empty        — the row carries no Links text

    Returns (buckets, ok). `ok` is False when the three do not sum to the vendor
    count, which is the only outcome that can hide a lost row.
    """
    edged = {e["vendor_ref"] for e in edges}
    flagged = {a["vendor_ref"] for a in ambiguous}
    with_links = {v[1] for v in vendors}
    buckets = {
        "total": len(all_vendors),
        "with_links": len(with_links),
        "empty": [r for r, has in all_vendors if not has],
        "edged": sorted(edged),
        "review_only": sorted(flagged - edged),
    }
    # A row that carries text but produced neither an edge nor a review item would
    # be the actual bug this accounting exists to catch. Named, not assumed.
    buckets["silent"] = sorted(with_links - edged - flagged)
    total = (len(buckets["edged"]) + len(buckets["review_only"])
             + len(buckets["empty"]) + len(buckets["silent"]))
    return buckets, total == buckets["total"]


def xlsx_crosscheck(path, db_labels):
    """[ORDER 32] Read the generated sheet's Links column and diff it against the
    shim column the parser reads. READ-ONLY, and never an input to the parse.

    Returns (rows_in_file, only_in_file, only_in_db, differing) or None when the
    file or openpyxl is unavailable (a missing sheet is a skipped check, not a
    failed run — the DB is the record).
    """
    try:
        import openpyxl
    except ImportError:
        return None
    if not Path(path).exists():
        return None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    file_labels = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = None
        for r in rows:
            cells = [str(c).strip() if c is not None else "" for c in r]
            if "ID" in cells and "Links" in cells:
                header = cells
                break
        if not header:
            continue
        i_id, i_lk = header.index("ID"), header.index("Links")
        for r in rows:
            ref = str(r[i_id]).strip() if i_id < len(r) and r[i_id] else ""
            val = str(r[i_lk]).strip() if i_lk < len(r) and r[i_lk] else ""
            if ref and val:
                file_labels[ref] = val
    wb.close()
    only_file = sorted(set(file_labels) - set(db_labels))
    only_db = sorted(set(db_labels) - set(file_labels))
    differing = sorted(r for r in set(file_labels) & set(db_labels)
                       if file_labels[r] != db_labels[r])
    return len(file_labels), only_file, only_db, differing


def write_report(path, all_vendors, vendors, edges, ambiguous, existing, reviewed,
                 applied, wrote, events, buckets, sums, xcheck, resolved_refs):
    mech = [e for e in edges if e["how"] == "mechanical"]
    namd = [e for e in edges if e["how"] == "name-matched"]
    dupes = [e for e in edges if (e["from_party"], e["to_party"], e["kind"]) in existing]
    L = []
    A = L.append
    A(f"# party_link parse — ORDER 17 / amendment 4 ({'APPLY' if applied else 'DRY RUN'})")
    A("")
    A(f"Generated {datetime.now(timezone.utc).isoformat(timespec='seconds')} · "
      f"{'wrote to the database' if applied else 'nothing written'}")
    A("")
    A("## Row accounting (ORDER 32 done-test — these must sum)")
    A("")
    A(f"- vendor rows in the record: **{buckets['total']}**")
    A(f"- rows that produced at least one edge: **{len(buckets['edged'])}**")
    A(f"- rows that produced only review items: **{len(buckets['review_only'])}** "
      f"({', '.join(buckets['review_only']) or 'none'})")
    A(f"- rows with no Links text at all (explicitly empty): **{len(buckets['empty'])}**")
    A(f"- rows carrying text that produced NEITHER an edge nor a review item: "
      f"**{len(buckets['silent'])}** {'(' + ', '.join(buckets['silent']) + ')' if buckets['silent'] else ''}")
    A("")
    A(f"**{len(buckets['edged'])} + {len(buckets['review_only'])} + {len(buckets['empty'])}"
      f" + {len(buckets['silent'])} = {len(buckets['edged']) + len(buckets['review_only']) + len(buckets['empty']) + len(buckets['silent'])}"
      f" · vendor rows {buckets['total']} · **{'SUMS' if sums else 'DOES NOT SUM — a row is unaccounted for'}**")
    A("")
    A("## xlsx cross-check (`DNA/Network/vendors.xlsx` Links column)")
    A("")
    if xcheck is None:
        A("Skipped — no `--xlsx` path given, or the file/openpyxl was unavailable. "
          "The database is the record; the sheet is a rendering of it.")
    else:
        n_file, only_file, only_db, differing = xcheck
        A(f"- Links values in the sheet: **{n_file}** · in `links_label`: **{len(vendors)}**")
        A(f"- in the sheet but NOT in the record (would be invisible to the parser): "
          f"**{len(only_file)}** {'(' + ', '.join(only_file) + ')' if only_file else ''}")
        A(f"- in the record but not in the sheet (stale export): "
          f"**{len(only_db)}** {'(' + ', '.join(only_db) + ')' if only_db else ''}")
        A(f"- present in both but with different text: "
          f"**{len(differing)}** {'(' + ', '.join(differing) + ')' if differing else ''}")
    A("")
    A("## Counts")
    A("")
    A(f"- vendors carrying a non-blank `links_label`: **{len(vendors)}**")
    A(f"- **mechanical edges** (ref token resolved): **{len(mech)}**")
    A(f"- **name-matched edges** (exact unique party name): **{len(namd)}**")
    A(f"- **ambiguous items** (never guessed): **{len(ambiguous)}**")
    A(f"- parseable total (edges the DB should hold after apply): **{len(edges)}**")
    A(f"- of those, already present before this run: **{len(dupes)}**")
    A(f"- rows actually inserted this run: **{wrote}**")
    A(f"- review `event` rows written this run: **{events}** "
      f"(vendors already carrying one: {len([v for v in vendors if v[0] in reviewed])})")
    A("")
    A("## Edges")
    A("")
    A("| vendor | kind | → target | how | item |")
    A("|---|---|---|---|---|")
    for e in edges:
        A(f"| {e['vendor_ref']} {e['vendor_name']} | `{e['kind']}` | {e['target']} | "
          f"{e['how']} | {e['item']} |")
    A("")
    A("## THE REVIEW LIST — Joe's call, never guessed")
    A("")
    A("`status` is RESOLVED where a `link-parties` edge already leaves this vendor: "
      "Joe answered the question and the grammar simply cannot read prose. OPEN is "
      "the work that is actually left.")
    A("")
    if not ambiguous:
        A("None.")
    else:
        A("| status | vendor | unparsed item | why |")
        A("|---|---|---|---|")
        for a in ambiguous:
            st = "RESOLVED" if a["vendor_ref"] in resolved_refs else "**OPEN**"
            A(f"| {st} | {a['vendor_ref']} {a['vendor_name']} | {a['item']} | {a['reason']} |")
        A("")
        n_open = len([a for a in ambiguous if a["vendor_ref"] not in resolved_refs])
        A(f"**{n_open} open · {len(ambiguous) - n_open} already resolved by a verb-written edge.**")
    A("")
    path.write_text("\n".join(L) + "\n")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true",
                    help="write the edges and the review events (default is a dry run)")
    ap.add_argument("--xlsx", default=None,
                    help="path to DNA/Network/vendors.xlsx — READ-ONLY drift check of "
                         "its Links column against links_label (never a parse input)")
    a = ap.parse_args()

    url = os.environ.get("DATABASE_URL") or os.environ.get("CARR_IMPORT_DB_URL")
    if not url:
        print("DATABASE_URL (or CARR_IMPORT_DB_URL) is not set — nothing attempted.",
              file=sys.stderr)
        return 78

    OUT.mkdir(exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    report_path = OUT / f"party-link-parse-{stamp}.md"

    with psycopg.connect(url) as conn:
        (vendors, all_vendors, refs, names, party_name,
         existing, reviewed, stated_from) = load(conn)
        edges, ambiguous = parse(vendors, refs, names, party_name)

        wrote = 0
        events = 0
        if a.apply:
            actor = conn.execute("select id from actor where slug='system'").fetchone()[0]
            for e in edges:
                key = (e["from_party"], e["to_party"], e["kind"])
                if key in existing:
                    continue
                conn.execute(
                    "insert into party_link (from_party, to_party, kind, note, source, created_by) "
                    "values (%s,%s,%s,%s,%s,%s)",
                    (e["from_party"], e["to_party"], e["kind"], e["note"],
                     PROVENANCE_SOURCE, actor))
                existing.add(key)
                wrote += 1

            by_vendor = {}
            for item in ambiguous:
                by_vendor.setdefault(item["vendor_ref"], []).append(item)
            vid_for = {v[1]: (v[0], v[5]) for v in vendors}
            for vref, items in by_vendor.items():
                vid, label = vid_for[vref]
                if vid in reviewed:
                    continue
                remainder = " · ".join(f"{i['item']}  [{i['reason']}]" for i in items)
                conn.execute(
                    "insert into event (occurred_at, actor_id, verb, subject_type, subject_id, "
                    "field, old_value, cause, agent_rationale) "
                    "values (now(), %s, %s, 'vendor', %s, 'links_label', to_jsonb(%s::text), "
                    "'import_migration', %s)",
                    (actor, REVIEW_VERB, vid, label,
                     "amendment-4 review: links_label text did not resolve under the "
                     "amendment-4 grammar and was NOT guessed — " + remainder +
                     ". Resolve by naming the party (then use link-parties), or confirm "
                     "the text is not a link."))
                reviewed.add(vid)
                events += 1
            conn.commit()

        buckets, sums = row_accounting(all_vendors, vendors, edges, ambiguous)
        resolved_refs = {v[1] for v in vendors if v[2] in stated_from}
        xcheck = xlsx_crosscheck(a.xlsx, {v[1]: (v[5] or "").strip()
                                          for v in vendors}) if a.xlsx else None

        write_report(report_path, all_vendors, vendors, edges, ambiguous, existing,
                     reviewed, a.apply, wrote, events, buckets, sums, xcheck,
                     resolved_refs)

    mech = sum(1 for e in edges if e["how"] == "mechanical")
    namd = sum(1 for e in edges if e["how"] == "name-matched")
    print(f"{'APPLIED' if a.apply else 'DRY RUN'} — vendors with labels {len(vendors)} · "
          f"mechanical {mech} · name-matched {namd} · ambiguous {len(ambiguous)} · "
          f"inserted {wrote} · review events {events}")
    print(f"row accounting — edged {len(buckets['edged'])} + review-only "
          f"{len(buckets['review_only'])} + empty {len(buckets['empty'])} + silent "
          f"{len(buckets['silent'])} = {buckets['total']} vendor rows · "
          f"{'SUMS' if sums else 'DOES NOT SUM'}")
    if xcheck is not None:
        n_file, only_file, only_db, differing = xcheck
        print(f"xlsx cross-check — sheet {n_file} · only-in-sheet {len(only_file)} · "
              f"only-in-record {len(only_db)} · text differs {len(differing)}")
    print(f"report: {report_path}")
    # An accounting that does not sum is a failed run, not a footnote.
    return 0 if sums else 1


if __name__ == "__main__":
    sys.exit(main())
