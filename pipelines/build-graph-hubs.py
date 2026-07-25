#!/usr/bin/env python3
"""
build-graph-hubs.py — connectivity layer for the PEOPLE graph.

THE PROBLEM IT SOLVES
  build-graph-notes.py draws an edge only when a referral field EXACTLY matches a
  node title, or a Links cell holds a literal V-### id. Measured 2026-07-25 that
  produced 105 edges across 663 nodes — 568 nodes (86%) completely isolated — while
  260 further references were dropped as "(no exact match)". More relationships
  were discarded than drawn.

  The exact-match rule is CORRECT as a record-safety rule (a lone surname must
  never merge two records) but it was the ONLY edge rule, so anything written as
  free text — "Kavin", "Jason Togni", "CARR Website", "Luke at Schein" — connected
  to nothing.

WHAT THIS ADDS
  Hub nodes for the things people have in COMMON, and links from each hub to its
  members: firm, city/territory, owner, category, and referral channel. A hub is
  an attribute, never a claim that two people are the same person, so it adds
  connectivity without ever risking a merge.

NON-DESTRUCTIVE BY DESIGN
  Writes ONLY to Graph/hubs/. It never edits the entity notes build-graph-notes.py
  produces, so it cannot corrupt them and cannot break that generator. Obsidian
  renders graph edges undirected, so hub -> member links cluster the graph without
  the member notes changing at all.

  Run AFTER `run.sh graph` (which wipes and rebuilds Graph/). `run.sh graph-hubs`
  does both in order.
"""
import sys, os, re, json, shutil, unicodedata
from collections import defaultdict
import openpyxl

ROOT = sys.argv[1] if len(sys.argv) > 1 else os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", ".."))
GRAPH = os.path.join(ROOT, "Graph")
OUT = os.path.join(GRAPH, "hubs")

VENDORS = os.path.join(ROOT, "DNA/Network/vendors.xlsx")
LEADS   = os.path.join(ROOT, "DNA/Leads/lead-registry.xlsx")
CLIENTS = os.path.join(ROOT, "DNA/Clients/client-roster.xlsx")
DEALS   = os.path.join(ROOT, "DNA/Deal Management/panhandle-team-deals.json")

def s(v): return str(v if v is not None else "").strip()

def rows(path, sheet):
    if not os.path.exists(path): return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames: wb.close(); return []
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = [s(h) for h in next(it, [])]
    out = [dict(zip(hdr, r)) for r in it if any(x is not None for x in r)]
    wb.close()
    return out

# ---------- learn the node titles build-graph-notes.py actually wrote ----------
node_titles = {}   # lowercase display name -> exact note title
for sub in ("vendors", "leads", "clients", "deals"):
    d = os.path.join(GRAPH, sub)
    if not os.path.isdir(d): continue
    for fn in os.listdir(d):
        if fn.endswith(".md"):
            node_titles[os.path.splitext(fn)[0].lower()] = os.path.splitext(fn)[0]

def node_for(*names, kind=None):
    """Resolve a person/entity to the exact note title, or None.

    build-graph-notes.py suffixes a lead/deal title with its record type when it
    would otherwise collide with the client or vendor of the same name (e.g.
    "Joseph Finelli" the client and "Joseph Finelli (deal)"). Try the suffixed
    form FIRST for those record types, otherwise every deal for an existing
    client would hub against the client node and the deal node would be left
    with no edges at all.
    """
    for n in names:
        n = s(n)
        if not n:
            continue
        if kind:
            k = f"{n} ({kind})".lower()
            if k in node_titles:
                return node_titles[k]
        if n.lower() in node_titles:
            return node_titles[n.lower()]
    return None

# ---------- normalise hub keys ----------
NOISE = {"", "n/a", "na", "none", "tbd", "(tbd)", "unknown", "-", "—", "unassigned",
         "(enrich)", "other", "other/not recorded"}

def norm_firm(v):
    t = s(v)
    if t.lower() in NOISE: return ""
    t = re.sub(r"[,.]?\s*(inc|llc|pllc|pa|pc|p\.a\.|l\.l\.c\.)\.?$", "", t, flags=re.I).strip()
    t = re.sub(r"\s+", " ", t)
    return t

def norm_city(v):
    t = s(v)
    if t.lower() in NOISE: return ""
    # "Destin / Panama City" and "Marietta | Smyrna" -> take each side
    return t

def split_multi(t):
    return [p.strip() for p in re.split(r"[/|;]| and ", t) if p.strip()]

# ---------- collect memberships ----------
hubs = defaultdict(lambda: defaultdict(set))   # kind -> hub name -> {node titles}

JUNK_RE = re.compile(r"tbd|enrich|not recorded|^\(|^\?+$", re.I)

def add(kind, key, node):
    """A hub key must be a real attribute value, not a placeholder."""
    if not key or not node: return
    k = key.strip()
    if k.lower() in NOISE or JUNK_RE.search(k) or len(k) < 2: return
    hubs[kind][k].add(node)

for v in rows(VENDORS, "Vendors"):
    n = node_for(v.get("Name"), v.get("Company"))
    if not n: continue
    add("Firm", norm_firm(v.get("Company")), n)
    for c in split_multi(norm_city(v.get("Territory"))): add("Market", c, n)
    add("Owner", s(v.get("Owner")) or "Unassigned", n)
    add("Category", s(v.get("Category")), n)

for l in rows(LEADS, "Registry"):
    n = node_for(l.get("Contact Name"), l.get("Practice"), kind="lead")
    if not n: continue
    add("Firm", norm_firm(l.get("Practice")), n)
    for c in split_multi(norm_city(l.get("City/Market"))): add("Market", c, n)
    add("Owner", s(l.get("Owner")) or "Unassigned", n)
    add("Specialty", s(l.get("Specialty")), n)
    add("Channel", s(l.get("Source Type")), n)
    # the free-text referrer that exact-match could never resolve
    det = s(l.get("Source Detail (V-ID / event / referrer)"))
    if det and not re.match(r"^V-[A-Z]+-\d+$", det.upper()) and len(det) < 60:
        add("Referrer", det, n)

for c in rows(CLIENTS, "Clients"):
    n = node_for(c.get("Name"), c.get("Practice / Entity"))
    if not n: continue
    add("Firm", norm_firm(c.get("Practice / Entity")), n)
    for m in split_multi(norm_city(c.get("Market / Location"))): add("Market", m, n)
    add("Owner", s(c.get("Owner")) or "Unassigned", n)
    add("Specialty", s(c.get("Specialty / Type")), n)
    ref = s(c.get("Referral Source"))
    if ref and len(ref) < 60: add("Referrer", ref, n)

if os.path.exists(DEALS):
    dj = json.load(open(DEALS))
    for d in dj.get("deals", []):
        n = node_for(d.get("name"), d.get("contact"), d.get("company"), kind="deal")
        if not n: continue
        for m in split_multi(norm_city(d.get("city"))): add("Market", m, n)
        add("Owner", s(d.get("owner")) or "Unassigned", n)
        ref = s(d.get("referral"))
        if ref and len(ref) < 60: add("Referrer", ref, n)
        add("Lane", s(d.get("lane")), n)

# ---------- write hubs ----------
if os.path.isdir(OUT): shutil.rmtree(OUT)
os.makedirs(OUT, exist_ok=True)

KIND_TAG = {"Firm": "hub-firm", "Market": "hub-market", "Owner": "hub-owner",
            "Category": "hub-category", "Specialty": "hub-specialty",
            "Channel": "hub-channel", "Referrer": "hub-referrer", "Lane": "hub-lane"}

def safe(name):
    x = unicodedata.normalize("NFKD", name)
    x = re.sub(r'[\\/:*?"<>|#^\[\]{}]', "", x).strip()
    return re.sub(r"\s+", " ", x) or "unnamed"

# ---------- what NOT to hub ----------
# Two rules, both learned by getting it wrong (Joe, 2026-07-25: "this thing is a
# giant hairball").
#
# 1. AN ATTRIBUTE THAT IS ALREADY A COLOUR MUST NOT ALSO BE AN EDGE. Owner is
#    rendered as colour (#owner-joe / #owner-dell) and Lane as the deal-* tags.
#    Hubbing them too added 533 edges — a third of all hub edges — that separate
#    nothing, because "Dell owns it" is true of 412 records. Two mega-stars
#    dominated the layout and drowned the informative structure.
# 2. A HUB LARGER THAN ~30 IS A TAXONOMY BUCKET, NOT A CLUSTER. "Banker/Lender"
#    with 47 members tells you the category system exists, which you knew. The
#    hubs that carry information are small: Firm averages 3 members, Referrer 5.
#
# Consequence, and it is the right one: records whose ONLY connection was their
# owner go back to being isolated. That is honest — they were never meaningfully
# connected, and their colour already says whose they are.
# 2. A hub is only a problem when it DOMINATES the layout. Obsidian's force
#    layout is fine with a 68-member market cluster in a 662-node graph (10%);
#    it collapses around a 412-member star (62%). So the cap scales with the
#    graph rather than being a flat number: anything pulling in more than ~12%
#    of all nodes is a taxonomy bucket, not a cluster. A flat cap of 30 was
#    tried first and cut too deep — connectivity fell to 65% and the orphan ring
#    came back.
SKIP_KINDS = {"Owner", "Lane"}
MAX_MEMBERS = max(30, int(0.12 * len(node_titles)))

written = 0
edges = 0
singletons = 0
skipped_big = []
taken = set()
for kind, group in hubs.items():
    if kind in SKIP_KINDS:
        continue
    for name, members in sorted(group.items()):
        if len(members) < 2:
            singletons += 1          # a hub of one connects nothing — skip it
            continue
        if len(members) > MAX_MEMBERS:
            skipped_big.append((len(members), kind, name))
            continue
        title = safe(f"{name} ({kind})")
        low = title.lower()
        if low in taken: continue
        taken.add(low)
        body = [
            "---", "type: hub", f'hub_kind: "{kind}"', f'hub_name: "{name}"',
            f"members: {len(members)}",
            f"tags: [hub, {KIND_TAG.get(kind,'hub-other')}]", "---", "",
            f"# {name}", "", f"*{kind} hub — {len(members)} records share this attribute.*",
            "", "## Members", "",
        ] + [f"- [[{m}]]" for m in sorted(members)] + [""]
        open(os.path.join(OUT, f"{title}.md"), "w", encoding="utf-8").write("\n".join(body))
        written += 1
        edges += len(members)

open(os.path.join(OUT, "README.md"), "w", encoding="utf-8").write(
    "# Graph/hubs — connectivity layer (DERIVED, DO NOT HAND-EDIT)\n\n"
    "Generated by carr-system/pipelines/build-graph-hubs.py AFTER build-graph-notes.py.\n"
    "A hub groups records that share an attribute (firm, market, owner, category,\n"
    "specialty, channel, referrer, lane). Hubs add graph connectivity WITHOUT ever\n"
    "asserting that two records are the same person — that remains the exact-match\n"
    "rule's job in the entity notes.\n\n"
    "Hubs with only one member are skipped: they connect nothing.\n\n"
    "Regenerate with `~/carr-system/run.sh graph-hubs` (runs graph, then hubs).\n")

# No silent caps: say exactly what was dropped and why.
if skipped_big:
    print(f"\nskipped {len(skipped_big)} hubs over {MAX_MEMBERS} members (taxonomy buckets, not clusters):")
    for n, kind, name in sorted(skipped_big, reverse=True):
        print(f"   {n:4d}  {kind}: {name}")
print(f"\nhubs: {written} written, {edges} edges, {singletons} single-member skipped, "
      f"{len(skipped_big)} oversized skipped, kinds excluded: {', '.join(sorted(SKIP_KINDS))} (rendered as colour instead)")
