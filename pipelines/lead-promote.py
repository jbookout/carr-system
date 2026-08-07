#!/usr/bin/env python3
"""
lead-promote.py — the promotion gate between the reservoir and the registry.

THE PROBLEM (found 2026-07-25)
  `DNA/Leads/lead-router-*.xlsx` holds 9,320 licensed providers in territory,
  segmented by buying signal. `lead-registry.xlsx` holds 199 worked rows. Only 37
  had ever crossed over, and 3 of 9,855 board records carry an agent's name.

  The cause is structural, not neglect: lead-system.md's intake architecture was
  designed 2026-07-06 and lists six sources — vendor intros, walk-ins, past-client
  referrals, Axon, the FL license feed, "anything else". The router file arrived a
  week later (2026-07-13) and was wired to the BOARD as a display layer. No intake
  SOP, no feed-router row, no promotion path. So 268 practices flagged ripe for
  sale and 228 in a lease-decision window sit where nothing can act on them.

WHAT THIS DOES — AND DELIBERATELY DOES NOT DO
  It does NOT bulk-import. 9,320 rows into a 199-row registry would drown the
  drip logic, the graph, the health checks and the Monday brief. The reservoir is
  a market map; the registry is a working pipeline. They should stay different sizes.

  It produces a weekly SHORTLIST from the event-driven segments, deduped against
  everything already known, ranked, ready for a human to claim.

  It does NOT write to the registry. Claim-before-touch is law (lead-system.md):
  "Nobody contacts a lead whose Owner is blank or Shared until an Owner gets
  stamped." And ownership is ASKED, never assumed (Joe's rule, 2026-07-13). So a
  row enters the registry only after a human claims it — this just decides who is
  worth asking about.

Usage:
  run.sh lead-promote [--count N] [--county NAME] [--segment KEY] [--all-segments] [--files]

SOURCE MODE (ORDER 29b, the ORDER 29a pattern). Reads default to RECORDS: the
router rows come from the record layer's pool (`v_export_pool` / the all-source
view), and lead-registry.xlsx / client-roster.xlsx / panhandle-team-deals.json
come from their own export views via lib/record_sources.py. --files forces the
historical read of the four generated files instead, and records mode falls
back to files, LOUDLY on stderr, whenever the pool or a view is unreachable —
never a silent short count.
"""
import sys, os, re, json, glob, argparse
from collections import Counter, defaultdict
from datetime import datetime, timezone
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    from lib.record_sources import (MODE_FILES, MODE_RECORDS, ROUTER_SOURCE, load_clients,
                                     load_deals_doc, load_leads, load_pool, pool_reach,
                                     resolve_mode, source_note)
    _HAVE_RECORDS = True
except ImportError:
    MODE_FILES, MODE_RECORDS, _HAVE_RECORDS = "files", "records", False

if _HAVE_RECORDS:
    _MODE, _rest = resolve_mode(sys.argv[1:], default=MODE_RECORDS)
else:
    _MODE = MODE_FILES
    _rest = [x for x in sys.argv[1:] if x not in ("--files", "--records")]
    if "--records" in sys.argv[1:]:
        print("[lead-promote] this copy has no lib/record_sources.py — running file mode",
              file=sys.stderr)

ap = argparse.ArgumentParser()
ap.add_argument("root", nargs="?")
ap.add_argument("--count", type=int, default=20, help="shortlist size (default 20)")
ap.add_argument("--county", help="restrict to one county")
ap.add_argument("--segment", help="restrict to one segment (substring match)")
ap.add_argument("--all-segments", action="store_true",
                help="include watch-list segments, not just event-driven ones")
a = ap.parse_args(_rest)

ROOT = a.root or os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# The renewal-radar lane (loop #204, Joe's ruling 2026-08-07). One of the
# LANE_SOURCES in lib/record_sources.py; named here so file mode can reach the
# same rows through the lane's own JSON when the record path is down.
RADAR_SOURCE = "renewal-radar"

MODE = _MODE
if MODE == MODE_RECORDS:
    _ok, _why, _, _ = pool_reach((ROUTER_SOURCE, RADAR_SOURCE))
    if not _ok:
        print(f"[lead-promote] records mode unavailable ({_why}) — falling back to the "
              f"generated files", file=sys.stderr)
        MODE = MODE_FILES

def s(v): return str(v if v is not None else "").strip()

def rows(path, sheet=None):
    if not os.path.exists(path): return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [s(h) for h in next(it, [])]
    out = [dict(zip(hdr, r)) for r in it if any(x is not None for x in r)]
    wb.close(); return out

# ---------- sources ----------
routers = sorted(glob.glob(os.path.join(ROOT, "DNA/Leads/lead-router-*.xlsx")))
if not routers:
    sys.exit("no lead-router-*.xlsx found in DNA/Leads/")
RESERVOIR = routers[-1]                                    # latest by name; the FILENAME
                                                             # still carries the router's own
                                                             # date for display, in both modes
if MODE == MODE_RECORDS:
    _pool = load_pool((ROUTER_SOURCE, RADAR_SOURCE))
    reservoir = _pool[ROUTER_SOURCE]
    radar_raw = _pool.get(RADAR_SOURCE, [])
    registry  = load_leads(ROOT, MODE_RECORDS)
    clients   = load_clients(ROOT, MODE_RECORDS)
    deals     = load_deals_doc(ROOT, MODE_RECORDS).get("deals", [])
else:
    reservoir = rows(RESERVOIR, "Lead Router")
    radarp    = os.path.join(ROOT, "Automation/renewal-radar.json")
    radar_raw = json.load(open(radarp)) if os.path.exists(radarp) else []
    registry  = rows(os.path.join(ROOT, "DNA/Leads/lead-registry.xlsx"), "Registry")
    clients   = rows(os.path.join(ROOT, "DNA/Clients/client-roster.xlsx"), "Clients")
    dealsp    = os.path.join(ROOT, "DNA/Deal Management/panhandle-team-deals.json")
    deals     = json.load(open(dealsp)).get("deals", []) if os.path.exists(dealsp) else []

print(f"[lead-promote] source: "
      f"{source_note(MODE) if _HAVE_RECORDS else 'generated files (no lib/record_sources.py)'}",
      file=sys.stderr)

# ---------- the dedupe gate ----------
# The lead-sweep workflow once created L-170 for a practice already a client at
# Closing. Anything promoted has to be checked against every record we hold, by
# BOTH name and email — name alone misses the spelling drift this vault is full of
# (Brielmayer/Breilmayer, Lindsay/Lindsey, Connor/Conner).
def norm(x):
    return re.sub(r"[^a-z]", "", s(x).lower())

known_names, known_emails = set(), set()
for r in registry:
    known_names.add(norm(r.get("Contact Name"))); known_names.add(norm(r.get("Practice")))
    known_emails.add(s(r.get("Email")).lower())
for c in clients:
    known_names.add(norm(c.get("Name"))); known_names.add(norm(c.get("Practice / Entity")))
    known_emails.add(s(c.get("Email")).lower())
for d in deals:
    for k in ("name", "contact", "company"):
        known_names.add(norm(d.get(k)))
    known_emails.add(s(d.get("email")).lower())
known_names.discard(""); known_emails.discard("")

# ---------- segments: THREE lanes, not one ----------
# Read from THE PLAY column rather than guessed. The segments do not all mean
# "work this lead" — they encode three different actions, and conflating them was
# the first draft's mistake.
LANES = {
    # LANE 1 — PROMOTE: a real 1:1 opportunity, becomes an L-### row on claim
    "POST-SALE FOUNDER":  ("PROMOTE", "Re-engagement",
        "sold, serving an earn-out; restarts with capital and a book — 'THE HOTTEST CLASS'"),
    "PRACTICE OWNER":     ("PROMOTE", "Re-engagement",
        "owns the entity: expansion, 2nd location, buy-vs-lease, refi"),
    "LEASE EVENT":        ("PROMOTE", "Renewal-Relocation", "lease decision window"),
    "RELOCATING OWNER":   ("PROMOTE", "Hot – Relocator", "moving into territory"),
    "OWNER-OCCUPIER":     ("PROMOTE", "Hot – Startup", "second location"),
    "NEW ENTITY":         ("PROMOTE", "Hot – Startup", "corp filing just landed"),
    "NATIONAL ACCOUNT":   ("PROMOTE", "National Account", "multi-location / expanding"),
    # LANE 2 — DRIP: nurture, never 1:1. lead-system.md: "Associate – Nurture …
    # straight to Nurture (Drip) — monthly newsletter list."
    "ASSOCIATE":          ("DRIP", "Associate – Nurture",
        "2-14 yrs in, owns nothing — 'be the only broker he knows'"),
    "DSO ASSOCIATE":      ("DRIP", "DSO / DSO-track", "DSO burnout → independent"),
    # LANE 3 — REFER: not a CARR client yet. Goes to a practice broker; CARR earns
    # the referral fee and the buyer's real estate later.
    "RIPE FOR SALE":      ("REFER", "—",
        "owner 60+, no succession → sells externally; refer to practice brokers"),
    # LANE 4 — WATCH: no action. Stays in the reservoir.
    "SOLO":               ("WATCH", "—", "entity unconfirmed — needs a Sunbiz check"),
    "NEW GRAD":           ("WATCH", "—", "too new, revisit in ~3 years"),
    "WINDING DOWN":       ("WATCH", "—", "retiring, not restarting — courtesy only"),
    "UNCLASSIFIED":       ("WATCH", "—", "unsegmented"),
    "INSTITUTIONAL":      ("WATCH", "—", "watch the physicians, not the institution"),
    "PRE-ENTITY":         ("WATCH", "—", "upstream watch"),
}
LANE_ORDER = {"PROMOTE": 0, "DRIP": 1, "REFER": 2, "WATCH": 3}

# Longest key first: "DSO ASSOCIATE" must not be swallowed by "ASSOCIATE".
_KEYS = sorted(LANES, key=len, reverse=True)

def classify(seg):
    u = seg.upper()
    for key in _KEYS:
        if key in u:
            return LANES[key]
    return ("WATCH", "—", "")

def lane_of(seg):     return classify(seg)[0]
def registry_segment(seg): return classify(seg)[1]
def why(seg):         return classify(seg)[2]

# ---------- the renewal-radar lane (loop #204) ----------
# Joe's ruling, 2026-08-07: the radar's T1 rows QUEUE FOR HIS REVIEW on this
# shortlist. Nothing here promotes, claims, or writes a lead — same law as the
# router rows above: Joe qualifies, the system never does.
#
# Lane rows arrive VERBATIM in build-renewal-feed.py's short-key shape
# (s/n/e/ph/co/ci plus le/tier/conf) in BOTH modes — records mode hands back the
# source_row jsonb the lane mapper wrote, file mode reads the lane's own
# Automation/renewal-radar.json — so ONE adapter onto the router's header names
# lets every gate below (lanes, dedupe, contactability) run unchanged on both
# sources. The reverse of this mapping lives in build-lead-board.py:load_router.
def adapt_radar(x):
    return {"SEGMENT": s(x.get("s")), "Name": s(x.get("n")), "Email": s(x.get("e")),
            "Phone": s(x.get("ph")), "County": s(x.get("co")), "City": s(x.get("ci")),
            "Profession": s(x.get("pr")), "Practice Address": s(x.get("ad")),
            "THE PLAY": "",
            # lease-event facts ride along to the shortlist and the review artifact
            "le": s(x.get("le")), "tier": s(x.get("tier")), "conf": s(x.get("conf")),
            "_radar": True, "_flag": s(x.get("flag")), "_rep": s(x.get("rep"))}

radar_all = [adapt_radar(x) for x in radar_raw]
# T1 only: a decision window inside 12 months is worth a claim conversation now;
# T2/T3 stay on the board and age into T1 via the feed's tier recompute.
radar_t1 = [r for r in radar_all if r["tier"].startswith("T1")]

# The lane's flag field carries TWO different facts and they get opposite
# treatment. "already L-xxx ..." is the feed's registry suppressor (fuzzier than
# the name gate below, so it catches spelling drift first): a known lead, never
# re-promoted. The GCCMLS rows instead carry a provenance flag ("... not yet
# tenant-identified"): a BUILDING-level signal with no person attached yet,
# which is not promotable but is also not a duplicate of anything.
def radar_known(r):  return r["_flag"].startswith("already")
def radar_building(r): return "not yet tenant-identified" in r["_flag"]

radar_buildings = [r for r in radar_t1 if radar_building(r)]
radar_uncontactable = [r for r in radar_t1
                       if not radar_known(r) and not radar_building(r)
                       and not (r["Email"] or r["Phone"])]

# ---------- filter ----------
cands, dropped_dupe = [], 0
for r in list(reservoir) + radar_t1:
    if r.get("_flag") and radar_known(r):
        dropped_dupe += 1; continue
    if r.get("_flag") and radar_building(r):
        continue                       # counted and surfaced separately above
    seg = s(r.get("SEGMENT"))
    if not (a.all_segments or lane_of(seg) in ("PROMOTE", "DRIP", "REFER")): continue
    if a.segment and a.segment.upper() not in seg.upper(): continue
    if a.county and a.county.lower() != s(r.get("County")).lower(): continue
    if norm(r.get("Name")) in known_names or s(r.get("Email")).lower() in known_emails:
        dropped_dupe += 1; continue
    if not (s(r.get("Email")) or s(r.get("Phone"))):   # uncontactable
        continue
    cands.append(r)

cands.sort(key=lambda r: (LANE_ORDER[lane_of(s(r.get("SEGMENT")))],
                          -len(s(r.get("Email"))), s(r.get("County")), s(r.get("Name"))))

# ---------- report ----------
print(f"\nLEAD PROMOTION — shortlist from {os.path.basename(RESERVOIR)} + renewal radar")
print("=" * 78)
print(f"reservoir {len(reservoir):,} · renewal radar {len(radar_all):,} "
      f"(T1 {len(radar_t1):,}) · already known (deduped out) {dropped_dupe:,} · "
      f"eligible {len(cands):,} · showing {min(a.count, len(cands))}")
if not a.all_segments:
    print("scope: event-driven segments only — "
          "use --all-segments to include associates / new grads / watch-list")
print()
lanes = Counter(lane_of(s(r.get("SEGMENT"))) for r in cands)
for ln in ("PROMOTE", "DRIP", "REFER"):
    if not lanes.get(ln): continue
    print(f"\n  {ln}  ({lanes[ln]:,})")
    for seg, n in Counter(s(r.get("SEGMENT")) for r in cands
                          if lane_of(s(r.get("SEGMENT"))) == ln).most_common():
        print(f"     {n:5,}  {seg}")
        print(f"            {why(seg)}")

print("\n" + "=" * 78)
print("SHORTLIST — claim before contact. Owner is asked, never assumed.\n")
for i, r in enumerate(cands[:a.count], 1):
    print(f"{i:3d}. {s(r.get('Name')):26s} {s(r.get('Profession'))[:22]:22s} "
          f"{s(r.get('City')):16s} {s(r.get('County'))}")
    print(f"     {s(r.get('SEGMENT'))}")
    if s(r.get("THE PLAY")): print(f"     play: {s(r.get('THE PLAY'))[:90]}")
    bits = [b for b in (s(r.get("Email")), s(r.get("Phone"))) if b]
    print(f"     {' · '.join(bits)}")
    if r.get("_radar"):
        facts = [b for b in (f"lease event {s(r.get('le'))}" if s(r.get("le")) else "",
                             s(r.get("tier")),
                             f"confidence {s(r.get('conf'))}" if s(r.get("conf")) else "",
                             s(r.get("_rep"))) if b]
        if facts: print(f"     {' · '.join(facts)}")
        print(f"     → registry Segment = {registry_segment(s(r.get('SEGMENT')))} · "
              f"Source Type = Renewal Radar · Source Detail = renewal-radar.json")
    else:
        print(f"     → registry Segment = {registry_segment(s(r.get('SEGMENT')))} · "
              f"Source Type = Lead Router · Source Detail = {os.path.basename(RESERVOIR)}")
    print()

print("=" * 78)
print("NEXT: Joe or Dell claims the ones they want. Only claimed rows become L-###")
print("rows — an unclaimed row stays in the reservoir rather than becoming an")
print("orphan with no owner (the 37 unowned sweep leads are exactly that failure).")
if radar_uncontactable:
    print(f"\nRENEWAL RADAR: {len(radar_uncontactable)} T1 decision-window rows carry no email "
          f"and no phone, so they cannot reach this shortlist yet. The fix is the DOH "
          f"licensure enrichment (DNA/Leads/renewal-radar-sop.md, step 7), which needs a "
          f"fresh download through Joe's authenticated browser — the raw file is never kept.")
if radar_buildings:
    print(f"\nRENEWAL RADAR: {len(radar_buildings)} further T1 rows are GCCMLS building-level "
          f"signals with no tenant identified yet — not promotable until someone names the "
          f"tenant, and listed in the review artifact so they stay visible.")

# ---------- the review artifact (loop #204) ----------
# Joe's ruling, 2026-08-07: T1 renewal candidates QUEUE FOR HIS REVIEW. This
# writes the full surviving T1 list (never capped at --count) where the brief
# pack reads it (pipelines/brief_pack.py, renewal-shortlist section), so the
# Monday brief presents it with Joe named as owner. out/ is gitignored, so the
# contact detail in here stays off git — same posture as the calendar archive.
# NOTHING here writes a lead: promotion stays a human claim at the board.
_shortdir = os.path.join(REPO, "out", "lead-promote")
os.makedirs(_shortdir, exist_ok=True)
_radar_cands = [r for r in cands if r.get("_radar")]
_artifact = {
    "built": datetime.now(timezone.utc).isoformat(),
    "source_note": source_note(MODE) if _HAVE_RECORDS else "generated files",
    "t1_total": len(radar_t1),
    "already_known": len([r for r in radar_t1 if radar_known(r)]),
    "uncontactable": len(radar_uncontactable),
    "building_signals": len(radar_buildings),
    # The feed's non-suppressor flag (a past-window or unparseable-date note) rides
    # along as `note`, so the review surface shows WHY a row needs a careful look.
    "candidates": [{**{k: v for k, v in r.items() if not k.startswith("_")},
                    **({"note": r["_flag"]} if r["_flag"] else {})}
                   for r in _radar_cands],
    "waiting_on_contact": [{"Name": r["Name"], "City": r["City"], "le": r["le"],
                            "tier": r["tier"], "conf": r["conf"]}
                           for r in radar_uncontactable],
    "buildings_no_tenant": [{"Name": r["Name"], "City": r["City"], "le": r["le"],
                             "tier": r["tier"]} for r in radar_buildings],
}
_shortpath = os.path.join(_shortdir, "renewal-t1-shortlist.json")
with open(_shortpath, "w") as fh:
    json.dump(_artifact, fh, indent=1)
print(f"\nreview artifact -> {_shortpath} ({len(_radar_cands)} candidates, "
      f"{len(radar_uncontactable)} waiting on contact info)")
