"""The five Wave 1 export targets (specs: record-layer/exporter-specs-2026-07-30.md).

xlsx targets use the TEMPLATE approach: the live workbook is opened as the
template, its static sheets (Config, Legend, Dashboard formulas, Intake Log)
are preserved untouched, and only the data sheet rows are rewritten. This is
what makes Dashboard formulas keep self-deriving and legacy consumers keep
parsing. At freeze, a frozen template copy replaces the live file as the
template source (exporters/templates/) so live-file drift can't leak in.

Graph nodes are NOT a separate exporter: run.sh graph already derives them
from these files, so the nightly order is: five exporters, then graph.
"""

import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment

from .common import VAULT, coverage_findings, coverage_note_cell, coverage_note_md
from .partial import splice
from .dictionary import DICT_REL, build_dictionary
from .ledger_targets import (HUNT_REL as LEDGER_HUNT_REL, RECIP_REL as LEDGER_RECIP_REL,
                             build_hunt_ledger, build_reciprocity)

TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"


def _template(rel):
    """Frozen template if present (post-freeze), else the live file (pre-cutover rehearsal)."""
    frozen = TEMPLATE_DIR / Path(rel).name
    return frozen if frozen.exists() else VAULT / rel


def _rewrite_sheet(wb, sheet_name, header, rows):
    ws = wb[sheet_name]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    live_header = [c.value for c in ws[1]][: len(header)]
    if [h for h in live_header if h] != [h for h in header if h]:
        raise ValueError(f"{sheet_name} header changed: {live_header} != {header}")
    for r in rows:
        ws.append(list(r))
    return ws


# The absence-ambiguous columns, per render. The phrase is what a blank does NOT
# tell the reader; common.coverage_findings supplies the live numbers.
#
# Deliberately short lists. Every one of these is a column whose blank a reader
# converts into a fact about the person ("nobody has called them", "there is
# nothing to do here") rather than a fact about our capture. Columns whose blank
# means what it looks like (Notes, Rivalry Group) stay off, or the note becomes
# wallpaper and stops being read.
TOUCH_WATCH = {
    "Last Touch": "whether anyone has been in contact",
    "Next Step": "whether something is owed on this relationship",
    "Next Action": "whether something is owed on this relationship",
    "Next Action Date": "when the next action is due",
}


def _annotate_coverage(ws, header, rows, noun, watched=TOUCH_WATCH):
    """Hang the coverage note off the header cell of each short column.

    NOT a prose row under the table and NOT an extra sheet. These workbooks are
    parsed by the Dashboard formulas, by the graph builder and by Dell's side, and
    a sentence sitting inside the data range is a row those readers try to read as
    a vendor. A cell comment changes no value at all, so the canonical rows and the
    A8 checksum are byte-identical with it and without it, and it lands exactly
    where the misreading starts: the reader is already looking at that column.
    """
    for finding in coverage_findings(rows, header, watched):
        cell = ws.cell(row=1, column=header.index(finding[0]) + 1)
        cell.comment = Comment(coverage_note_cell(finding, noun), "CARR record layer")
        cell.comment.width = 340
        cell.comment.height = 130


# ---------------- lead-registry.xlsx ----------------

REGISTRY_REL = "DNA/Leads/lead-registry.xlsx"
REGISTRY_COLS = ["Lead ID", "Date In", "Owner", "Stage", "Segment", "Contact Name", "Practice",
                 "Specialty", "City/Market", "County", "Email", "Phone", "Source Type",
                 "Source Detail (V-ID / event / referrer)", "Report-Back Due", "Drip Campaign",
                 "Drip Added", "Next Action", "Next Action Date", "Last Touch", "SF Deal",
                 "Detail File", "Notes", "Est-Lease-Event", "Event-Source", "Event-Confidence"]


def build_registry(tmp_path, cur):
    cur.execute('select * from v_export_leads order by "Lead ID"')
    cols = [d[0] for d in cur.description]
    data = cur.fetchall()
    rows = [[r[cols.index(c)] for c in REGISTRY_COLS] for r in data]  # _suppressed carried in DB, not a sheet column
    wb = openpyxl.load_workbook(_template(REGISTRY_REL))
    ws = _rewrite_sheet(wb, "Registry", REGISTRY_COLS, rows)
    _annotate_coverage(ws, REGISTRY_COLS, rows, "leads")
    wb.save(tmp_path)
    return len(rows), rows


# ---------------- client-roster.xlsx ----------------

ROSTER_REL = "DNA/Clients/client-roster.xlsx"
ROSTER_COLS = ["Client ID", "Name", "Practice / Entity", "Owner", "Status", "Specialty / Type",
               "Market / Location", "Deal Type", "Referral Source", "Contact", "Phone", "Email",
               "Possible Duplicate Of", "Detail File", "Notes"]


def build_roster(tmp_path, cur):
    cur.execute('select * from v_export_clients order by "Client ID"')
    cols = [d[0] for d in cur.description]
    rows = [[r[cols.index(c)] for c in ROSTER_COLS] for r in cur.fetchall()]
    wb = openpyxl.load_workbook(_template(ROSTER_REL))
    _rewrite_sheet(wb, "Clients", ROSTER_COLS, rows)
    wb.save(tmp_path)
    return len(rows), rows


# ---------------- vendors.xlsx ----------------

VENDORS_REL = "DNA/Network/vendors.xlsx"
VENDORS_COLS = ["ID", "Name", "Company", "Category", "Vertical", "Title", "Owner", "Stage",
                "Last Touch", "Next Step", "Referral-active?", "Territory", "State", "Offers",
                "Seeking", "Links", "Rivalry Group", "Originated / Referred", "Phone", "Email",
                "Notes", "Enrich?"]


def build_vendors(tmp_path, cur):
    cur.execute('select * from v_export_vendors order by "ID"')
    cols = [d[0] for d in cur.description]
    data = cur.fetchall()
    in_market = [[r[cols.index(c)] for c in VENDORS_COLS]
                 for r in data if not r[cols.index("_out_of_market")]]
    wb = openpyxl.load_workbook(_template(VENDORS_REL))
    ws = _rewrite_sheet(wb, "Vendors", VENDORS_COLS, in_market)
    # Counted over the IN-MARKET rows, which are the rows this sheet holds. Counting
    # the whole view would state a coverage figure for rows the reader cannot see.
    _annotate_coverage(ws, VENDORS_COLS, in_market, "vendors")
    # Out of Market sheet keeps its explanatory first row; data rows follow it.
    # v1: out-of-market rows stay wherever the sheet held them at freeze; the
    # flag routes NEW moves. Revisit at freeze with real data.
    wb.save(tmp_path)
    return len(in_market), in_market


# ---------------- lead-router-2026-07-13.xlsx (target #8, Wave 3) ----------------
#
# The router regenerates from candidate_pool so every remaining reader keeps
# working — Dell's side included, and Dell has no DB path at all (ORDER 28's
# central finding). It rides the same A8 gate as the other seven.
#
# DEATH SENTENCE, recorded here and in the amendment-5 shim registry: this target
# retires at the Wave 4 repoint, once the board view is CONFIRMED the only reader
# — confirmed, not assumed. Until then it regenerates nightly.
#
# FIDELITY: the sheet's 17 columns split in two. Nine are DB-owned and come from
# the record. The other eight (Owns?, SUNBIZ entities, Lic Yrs, Licensed, Age
# Band, # at Address, Typical Term (est), License) pass back out of source_row
# verbatim with their native types intact — the same rule build_deals applies to
# the legacy deal fields. Row ORDER is the source file's, restored from
# source_seq: jsonb carries no order, and a reshuffled sheet is a diff nobody
# can read.
#
# EVERY POOL ROW EXPORTS, whatever its status. A suppressed_dup is still a row and
# a promoted row is still part of the market map. Filtering here would quietly
# shrink the file Dell reads, which is never-pre-qualify failing at the far end.

ROUTER_REL = "DNA/Leads/lead-router-2026-07-13.xlsx"
ROUTER_SHEET = "Lead Router"
ROUTER_COLS = ["SEGMENT", "THE PLAY", "Owns?", "SUNBIZ entities", "Name", "Profession",
               "Lic Yrs", "Licensed", "Age Band", "# at Address", "Practice Address",
               "City", "County", "Typical Term (est)", "Email", "Phone", "License"]
ROUTER_DB_OWNED = {
    "SEGMENT": "SEGMENT", "THE PLAY": "THE PLAY", "Name": "Name", "Profession": "Profession",
    "Practice Address": "Practice Address", "City": "City", "County": "County",
    "Email": "Email", "Phone": "Phone",
}


def build_router(tmp_path, cur):
    cur.execute("select * from v_export_pool order by source_seq")
    cols = [d[0] for d in cur.description]
    rows = []
    for r in cur.fetchall():
        rec = dict(zip(cols, r))
        legacy = rec.get("source_row") or {}
        rows.append([rec[ROUTER_DB_OWNED[c]] if c in ROUTER_DB_OWNED else legacy.get(c)
                     for c in ROUTER_COLS])
    wb = openpyxl.load_workbook(_template(ROUTER_REL))
    _rewrite_sheet(wb, ROUTER_SHEET, ROUTER_COLS, rows)
    wb.save(tmp_path)
    return len(rows), rows


# ---------------- panhandle-team-deals.json ----------------

DEALS_REL = "DNA/Deal Management/panhandle-team-deals.json"


def build_deals(tmp_path, cur):
    import json
    from datetime import datetime, timezone
    cur.execute("select * from v_export_deals order by name")
    cols = [d[0] for d in cur.description]
    deals = []
    for r in cur.fetchall():
        row = dict(zip(cols, r))
        legacy = row.get("source_row") or {}
        # FIDELITY RULE (reconciliation round 2): legacy passthrough wins for
        # vocabulary-rich fields (txn, carr) the DB normalizes internally;
        # DB values win for fields it actively owns (name, phase, owner, seg)
        # — those now render display-faithfully (phase label, initcap owner).
        legacy.update({
            "name": row["name"], "phase": row["phase"],
            "owner": row["owner"] or legacy.get("owner"),
            "txn": legacy["txn"] if "txn" in legacy else row["deal_type"],
            "seg": legacy["seg"] if "seg" in legacy else row["segment"],
            "carr": legacy.get("carr") if legacy.get("carr") is not None
                    else row["PLACEHOLDER_sf_commission_never_sum"],
        })
        deals.append(legacy)
    doc = {
        "source": "GENERATED from the CARR record layer — do not hand-edit; regenerated nightly",
        "captured": datetime.now(timezone.utc).isoformat(),
        "count": len(deals),
        "placeholders": "Salesforce Total Commission and Close Date are placeholders: never sum, never forecast, never rank by them.",
        "notes": "OPEN-pipeline view. This file was never the full deal record and still is not (see memory: Outlook deal folders are the real record).",
        "deals": deals,
        "schema": "see v_export_deals + record-layer/exporter-specs-2026-07-30.md",
    }
    tmp_path.write_text(json.dumps(doc, indent=2, default=str))
    return len(deals), deals


# ---------------- clients-active.md ----------------

ACTIVE_REL = "DNA/Clients/clients-active.md"
ACTIVE_COLS = ["Owner", "Name", "C-ID", "Status", "Deal Type", "Specialty", "Location",
               "Last Touch", "Next Step", "Detail"]


FROZEN_ACTIVE = Path(__file__).resolve().parent.parent / "frozen-sources" / "2026-07-30" / "clients-active.md"


# A COUNT FROZEN INTO PROSE IS A LIE ON A TIMER. The frozen header says "the full
# universe (117 owner-tagged records)". 117 was true on 2026-07-11 and has not been
# true since; the roster now carries 160. Nothing recounted it, because the sentence
# was carried across verbatim with the rest of the identity prose -- exactly the
# no-hardcoded-counts failure the coverage notes exist to prevent elsewhere in this
# file. The number is now recounted from the same view the roster is built from
# (v_export_clients), on every export, so the claim cannot drift from the file it
# describes. FAIL-SOFT ON PURPOSE: if the pattern is ever absent from the source
# prose, the header is emitted unchanged rather than the export dying over a
# sentence. A pattern that stops matching shows up as the stale number returning,
# which is visible; an exporter that refuses to run is not.
_UNIVERSE_COUNT_RE = re.compile(r"\(\s*[\d,]+\s+owner-tagged records\s*\)")


def _lifted_header(cur=None):
    """The file's own prose, carried from the frozen copy, with counts recounted.

    The identity of this file -- what it is, what it replaced, where narrative
    lives -- was written by Joe and is not the exporter's to paraphrase. Three
    things change: the hand-maintained 'Last updated' / 'Last synced' stamps are
    dropped, because they would be stale the moment they were copied into a
    nightly-regenerated file (no-fabrication applies to metadata too), and any
    "(N owner-tagged records)" claim is recounted live rather than carried.
    """
    keep = []
    for line in FROZEN_ACTIVE.read_text().splitlines():
        if line.startswith("## "):
            break
        if line.startswith("# ") or line.startswith("Last updated:") or line.startswith("Last synced:"):
            continue
        keep.append(line)
    while keep and not keep[-1].strip():
        keep.pop()

    if cur is not None:
        cur.execute("select count(*) from v_export_clients")
        universe = cur.fetchone()[0]
        keep = [_UNIVERSE_COUNT_RE.sub(f"({universe:,} owner-tagged records)", ln) for ln in keep]
    return keep


def _md_cell(v):
    """Render one value as a markdown table cell.

    A pipe inside a value silently splits the row into an extra column and every
    cell after it shifts left -- C-131's location ("Marietta | Smyrna") did exactly
    that, putting a city where the Next Step belonged. Newlines end the row
    outright. Both are escaped rather than stripped: the data stays verbatim, the
    table stays parseable.
    """
    if v is None:
        return ""
    return str(v).replace("\\", "\\\\").replace("|", "\\|").replace("\r", " ").replace("\n", "<br>")


def build_clients_active(tmp_path, cur):
    cur.execute('select * from v_export_clients_active order by "Owner", "Name"')
    cols = [d[0] for d in cur.description]
    rows = [[r[cols.index(c)] for c in ACTIVE_COLS] for r in cur.fetchall()]
    lines = [
        "# Clients — Shared Active Index (both partners, one book)",
        "",
        "> **GENERATED from the CARR record layer — do not hand-edit; regenerated nightly.**",
        "> Membership is DERIVED (an open deal, or a status flagged as pipeline-active), not",
        "> stored. Records change via the MCP verbs (log-activity, update-deal,",
        "> set-next-action...); this file is a rendered view. Where the prose below predates",
        "> the record layer and says to update rows in place, the MCP verbs are how you do it now.",
        "",
        *_lifted_header(cur),
        "",
        "## Active pipeline",
        "",
        "| " + " | ".join(ACTIVE_COLS) + " |",
        "|" + "---|" * len(ACTIVE_COLS),
    ]
    for r in rows:
        lines.append("| " + " | ".join(_md_cell(v) for v in r) + " |")
    lines += coverage_note_md(rows, ACTIVE_COLS, TOUCH_WATCH, "clients")
    from datetime import datetime, timezone
    lines += ["", f"*Exported: {datetime.now(timezone.utc).isoformat()}*", ""]
    tmp_path.write_text("\n".join(lines))
    return len(rows), rows


# ---------------- compiled rules (the taught-rules loop) ----------------

RULES_SHARED_REL = "DNA/compiled-rules-shared.md"
RULES_JOE_REL = "00_Context/compiled-rules-joe.md"
RULES_INTRO_REL = "DNA/Network/introduction-rules.md"
RULES_DELL_REL = "DNA/compiled-rules-dell.md"

# ORDER 37. introduction-rules.md is DOMAIN-scoped: vendor↔vendor politics, read
# by engine.md before it proposes an intro and before any joint guest list goes
# out. Those rules live in the SAME store as every other taught rule, and they
# are shared (personal_to null — Joe and Dell both teach them), so without a
# filter every one of them would ALSO print in compiled-rules-shared.md and the
# always-read core would silently gain nineteen vendor-politics rules nobody
# asked it to carry. The shared and personal renders therefore EXCLUDE this
# scope and the intro render INCLUDES only it: one store, three audiences, no
# rule printed in two files.
INTRO_SCOPE = "intro_politics"


def _is_intro(row):
    return (row.get("scope") or {}).get("kind") == INTRO_SCOPE


def _rules_header(scope_line):
    return [
        "> **GENERATED from the CARR record layer's rule store — do not hand-edit.**",
        "> These rules BIND like the standing rules in `CARR AI/CLAUDE.md`.",
        f"> {scope_line}",
        ">",
        "> **To add a rule, do not edit this file.** Capture it with the `teach` verb",
        "> (the human's verbatim words as `human_quote`), get the human's yes via",
        "> `activate-rule`, then refresh with `run.sh export --only compiled-rules`.",
        "> Only ACTIVE rules appear here — a proposed rule binds nobody by design.",
        "",
    ]


# ---------------- the rule taxonomy (PASS 1, 2026-08-03) ----------------
#
# WHY THIS EXISTS. compiled-rules-shared.md was a FLAT BULLET LIST: 28 rules in
# 35,991 bytes, roughly 1.3 KB per bullet. The curation plan
# (specs/rule-curation-plan-2026-08-02.md) takes shared from 29 rules to 73,
# which is about 94 KB of unbroken bullets in one always-read file. That breaks
# active rule 7e9739f2 ("clear, concise, easily scannable... no walls of prose")
# at the file level, and it makes the recitation active rule 4f7c348f requires —
# state the count per file, then name only the handful bearing on this task —
# impossible to perform, because selecting a handful needs a browsable file.
#
# The render below is modelled on build_rules_intro, which has been doing exactly
# this since ORDER 37: a DECLARED (key, heading) table, grouped output, and empty
# sections PRINTED rather than dropped. Same reasoning as there — a missing
# heading reads as "no such category exists", an empty one reads as "nothing is
# active here yet", and while 44 imported rules sit proposed and waiting on Joe,
# the second statement is the true one.
#
# WHERE A RULE'S SECTION COMES FROM, in precedence order:
#   1. `scope.section` on the rule row. Every one of the 52 rules imported from
#      ai-operating-notes.md carries it (the importer wrote the source heading
#      into scope), so the whole bulk backlog maps for free the moment it
#      activates. SECTION_ALIASES translates those source headings onto the
#      canonical keys below.
#   2. SECTION_OVERRIDES, matched on a statement prefix. A one-time bridge for
#      the 44 rules taught BEFORE this taxonomy existed and therefore carrying
#      no scope.section. See its own note.
#   3. UNCATEGORISED. Visible, named, and normally empty.
#
# THE DURABLE PATH IS (1): `teach` already accepts a free-form scope, so a rule
# taught with scope {"section": "<key>"} lands in its section with no code edit.
# The valid keys are printed in the file itself for exactly that reason.

UNCATEGORISED = "uncategorised"

RULE_SECTIONS = [
    ("core-conduct",         "Core conduct — how a session behaves"),
    ("authority-delegation", "Authority and delegation — who decides, who does the work"),
    ("communication",        "Talking to a partner — how output reaches a human"),
    ("partners",             "The partnership — the team, and reading each partner"),
    ("verification",         "Verification, research, and handling sources"),
    ("record-layer",         "The record layer — records, merges, and data integrity"),
    ("deal-work",            "Client and deal work"),
    ("carr-voice-business",  "CARR voice, positioning, and writing"),
    ("operating-mechanics",  "Operating mechanics — automation, gates, and source truth"),
    ("system-structure",     "System structure — files, folders, and scaffolding"),
    ("tools-platforms",      "Tools and platforms"),
    ("rule-loop",            "The rule loop itself"),
    (UNCATEGORISED,          "Uncategorised — no section recorded"),
]

RULE_SECTION_KEYS = [k for k, _ in RULE_SECTIONS]

# scope.section value (ai-operating-notes.md's own headings, as the importer
# slugged them) -> canonical key above. The two trailing-hyphen keys are the
# importer's 28-character truncation, verbatim; they are matched as stored, not
# tidied, because the stored value is what the row actually carries.
#
# `division-of-labor` maps to tools-platforms, not to partners: its single rule
# is "Copilot owns Microsoft-native, Claude owns everything else", which is a
# division of TOOL labour, not of the two humans' work.
# `channels-routing` maps to system-structure: two of its three rules retire as
# CLAUDE.md duplicates and the survivor is the artifact-tombstone rule.
# `standing-mechanics` keeps its own section (operating-mechanics) rather than
# being folded into system-structure — it holds autonomous-run requirements,
# risk colours and the ship gate, which are not file-and-folder rules.
SECTION_ALIASES = {
    "core-conduct":                 "core-conduct",
    "carr-voice-business":          "carr-voice-business",
    "division-of-labor":            "tools-platforms",
    "standing-mechanics":           "operating-mechanics",
    "file-folder-standards":        "system-structure",
    "scaffolding-policy-judgment-": "system-structure",
    "channels-routing":             "system-structure",
    "working-with-joe":             "partners",
    "the-partner-impact-test-jul-": "partners",
}

# THE BRIDGE, and it is deliberately a bridge and not a permanent home.
#
# The 44 rules active on 2026-08-03 were taught one at a time, before any
# taxonomy existed, so none carries scope.section. Dropping all 44 into
# UNCATEGORISED would have shipped a "sectioned" file that is one 36 KB bucket
# and twelve empty headings — the wall this pass exists to remove, wearing a
# table of contents. So they are placed here, in code, reviewably, keyed on a
# statement prefix (the view publishes no rule id — see the note on
# _rule_ident) with the prefix long enough to be unique across the store.
#
# A stale key is harmless: it matches nothing and the count check below reports
# it. A NEW rule taught without scope.section lands in UNCATEGORISED, which is
# the point — an empty Uncategorised section is the signal that every rule has a
# home, and a populated one is the signal that someone taught a rule without a
# section. Do not grow this table for new rules; pass scope.section to `teach`.
SECTION_OVERRIDES = {
    "On every new client, run open-source":     "verification",
    "Capture and rendering are separate":       "verification",
    "A claim about what a surface":             "verification",
    "Verification backfill":                    "verification",
    "When ANY new contact enters the system":   "verification",
    "When Joe provides an article, video":      "verification",
    "When Joe feeds a link or source":          "verification",
    "When Joe feeds a link (article":           "verification",
    "Every client search packet includes":      "deal-work",
    "Cold and paused clients leave":            "deal-work",
    "Outbound document formats":                "deal-work",
    "Every LOI term is negotiable":             "deal-work",
    "NATIONAL ACCOUNT MODELLING":               "deal-work",
    "Whenever a partner is asked to edit":      "communication",
    "When Joe needs to make several calls":     "communication",
    "All dialogue with Joe must be clear":      "communication",
    "CoStar is driven ONLY":                    "tools-platforms",
    "For any x.com/X link":                     "tools-platforms",
    "Partnership complementarity ledger":       "partners",
    "Dell experience calibration":              "partners",
    "Complementarity ledger data-parity":       "partners",
    "DELL IS A FULL WRITER":                    "partners",
    "Joe's strength-training kit":              "partners",
    "Joe's curriculum rule":                    "partners",
    "Joe's development kit":                    "partners",
    "Steelman Joe's labels":                    "partners",
    "Read Joe brokerage-first":                 "partners",
    "Joe's stated operating posture":           "partners",
    "Every CARR session carries the Dr":        "carr-voice-business",
    "WRITING-RULES BINDS PROSPECT-VISIBLE":     "carr-voice-business",
    "THE TEST FOR CAPTURING A TAUGHT RULE":     "rule-loop",
    "EVERY session recites BOTH rule files":    "rule-loop",
    "Before claiming the record layer cannot":  "record-layer",
    "Vendor relationship levels":               "record-layer",
    "Findings, enrichment results":             "record-layer",
    "A CARR AGENT'S OWN CONTACT DETAILS":       "record-layer",
    "SURVIVORSHIP RULE FOR DUPLICATE PARTIES":  "record-layer",
    "THE COO SEAT DECIDES AND REPORTS":         "authority-delegation",
    "THE STAFFING MODEL":                       "authority-delegation",
    "THE MAIN SESSION IS THE CEO":              "authority-delegation",
    "Do not stop for permission on mechanical": "authority-delegation",
    "BE THE SAFETY CHECK":                      "core-conduct",
    "On any non-trivial problem or decision":   "core-conduct",
    "Never tell Joe to delete a file":          "system-structure",
}


def _rule_section(row):
    """The canonical section key for one rule row. Never raises, never guesses."""
    stored = (row.get("scope") or {}).get("section")
    if stored:
        if stored in SECTION_ALIASES:
            return SECTION_ALIASES[stored]
        if stored in RULE_SECTION_KEYS:
            return stored                      # taught with a canonical key already
        # An unrecognised stored section is NOT silently remapped: a wrong home is
        # worse than a visible one. It surfaces in Uncategorised, named.
        return UNCATEGORISED
    statement = row.get("statement") or ""
    for prefix, key in SECTION_OVERRIDES.items():
        if statement.startswith(prefix):
            return key
    return UNCATEGORISED


def _rule_ident(row):
    """The rule's id, rendered as a trailing tag, or '' when it is not published.

    v_compiled_rules does NOT expose `rule.id` today (0015, widened by 0029 to add
    `scope`), and the exporter credential is denied the `rule` table on purpose —
    0029's own note explains why the view, not a wider grant, is the way to publish
    a column. So this returns nothing right now and no fake identifier is invented.
    The moment a migration appends `r.id` to the view, ids start printing here and
    in the CLAUDE.md gist index with no further code change.
    """
    rid = row.get("id")
    return f"  `#{str(rid)[:8]}`" if rid else ""


def _rule_when(row):
    return row["activated_at"].date().isoformat() if row["activated_at"] else "date unrecorded"


def _rule_bullet(row):
    line = f"- **{row['statement']}** — taught by {row['taught_by']}, {_rule_when(row)}"
    if row["human_quote"]:
        line += f' ("{row["human_quote"]}")'
    if row["enforcement"] and row["enforcement"] != "prose":
        line += f"  `[{row['enforcement']}]`"
    return line + _rule_ident(row)


def _fetch_rules(cur, personal_slug):
    """Active, non-intro rules for one audience, in a stable order.

    The view's own `order by activated_at` is not quite deterministic (two rules
    activated in the same transaction can tie), and a render that reshuffles
    between exports produces diffs that mean nothing. Statement breaks the tie.
    """
    if personal_slug is None:
        cur.execute("select * from v_compiled_rules where personal_to is null")
    else:
        cur.execute("select * from v_compiled_rules where personal_to = %s", (personal_slug,))
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    rows = [r for r in rows if not _is_intro(r)]        # ORDER 37, see INTRO_SCOPE
    rows.sort(key=lambda r: (r["activated_at"] or datetime.min.replace(tzinfo=timezone.utc),
                             r["statement"]))
    return rows


def _section_index_block(rows, noun):
    """The per-section count table that sits under the header.

    Active rule 4f7c348f makes every session recite the COUNT for each rule file
    and then name only the rules bearing on the task. Both halves of that are a
    lookup, and before this table both were a manual tally of a flat bullet list.
    The total is printed too, so the count a session recites is read, not counted.
    """
    counts = {}
    for r in rows:
        counts[_rule_section(r)] = counts.get(_rule_section(r), 0) + 1
    out = [f"**{len(rows)} active {noun}, by section.** Recite the total; read the section",
           "your task sits in.",
           "",
           "| Section | Rules |",
           "|---|---|"]
    for key, heading in RULE_SECTIONS:
        out.append(f"| {heading} | {counts.get(key, 0)} |")
    out += [f"| **Total** | **{len(rows)}** |", ""]
    return out


def _build_rules(tmp_path, cur, personal_slug, title, scope_line):
    """One bullet per active rule, grouped into the declared sections."""
    rows = _fetch_rules(cur, personal_slug)
    lines = [f"# {title}", ""] + _rules_header(scope_line)
    if not rows:
        # Thirteen empty headings under "no active rules yet" is thirteen ways of
        # saying the same nothing. compiled-rules-dell.md is empty by design until
        # Dell teaches his first rule, and that file is his first impression of
        # the mechanism.
        lines += ["*No active rules yet. The first one lands here the moment it is taught "
                  "and activated.*", ""]
    else:
        lines += _section_index_block(rows, "rule(s)")

    for key, heading in RULE_SECTIONS if rows else []:
        hits = [r for r in rows if _rule_section(r) == key]
        if key == UNCATEGORISED and not hits:
            # The ONE section that is dropped when empty, and only when empty.
            # Uncategorised is a defect report, not a category: printing "no
            # active rule in this section" under it every export trains the eye
            # to skip the heading that is supposed to catch attention the day a
            # rule does land there.
            continue
        lines += [f"## {heading}", ""]
        if not hits:
            lines += ["*No active rule in this section.*", ""]
            continue
        lines += [_rule_bullet(r) for r in hits]
        lines.append("")

    lines += [
        "---",
        "",
        "*Sections are declared in `exporters/targets.py` (`RULE_SECTIONS`). A rule lands",
        "in one by its stored `scope.section`; teach a new rule with",
        f"`scope: {{\"section\": \"...\"}}` using one of: "
        f"{', '.join(k for k in RULE_SECTION_KEYS if k != UNCATEGORISED)}.",
        "A rule taught without one appears under Uncategorised, which is normally absent.*",
        "",
    ]
    lines += [f"*Exported: {datetime.now(timezone.utc).isoformat()} · "
              f"{len(rows)} active rule(s)*", ""]
    tmp_path.write_text("\n".join(lines))
    # canonical rows for the checksum: the rule content and its placement, not the
    # timestamp line. Placement is in because a rule silently changing section is a
    # content change to this file and the checksum is what notices content changes.
    return len(rows), [[r["statement"], r["human_quote"], r["taught_by"],
                        r["personal_to"], r["enforcement"], _rule_section(r)] for r in rows]


def build_rules_shared(tmp_path, cur):
    return _build_rules(
        tmp_path, cur, None, "Compiled rules — SHARED (both partners)",
        "SHARED SCOPE: these apply to Joe's brain and Dell's brain alike.")


def build_rules_joe(tmp_path, cur):
    return _build_rules(
        tmp_path, cur, "joe", "Compiled rules — Joe (personal)",
        "PERSONAL SCOPE: these apply to Joe's brain only. Dell's equivalent file "
        "is DNA/compiled-rules-dell.md, generated the same way.")


def build_rules_dell(tmp_path, cur):
    # Added 2026-08-01 (Dell onboarding prep). Renders into the DNA share — the
    # one surface Dell's Mac reaches — so his sessions auto-load it exactly as
    # Joe's load compiled-rules-joe. Empty until Dell teaches his first rule;
    # an empty file with the header is the correct day-one state, not a bug.
    return _build_rules(
        tmp_path, cur, "dell", "Compiled rules — Dell (personal)",
        "PERSONAL SCOPE: these apply to Dell's brain only. Joe's equivalent is "
        "00_Context/compiled-rules-joe.md on his side.")


# -------- the rule-gist index inside CLAUDE.md (PASS 1, 2026-08-03) --------
#
# JOE'S RULING, 2026-08-02: the rule store holds all binding text with its scope
# and provenance; CLAUDE.md carries a ONE-LINE GIST per rule naming the store as
# the authority; and that gist list is GENERATED, because a hand-written summary
# drifts from the rule it summarises — the same two-homes disease one layer down.
#
# WHICH CLAUDE.md, and why this one. There are two:
#   * `My Drive/.claude/CLAUDE.md` — Claude Code's standing context, and by its
#     own header a deliberate STANDALONE FALLBACK, "useful standalone, even if
#     the CARR AI Drive folder isn't connected". A generated block there could
#     only be refreshed when the vault IS reachable, so it would be stale in
#     exactly the situation that file exists for. It is also Mac-local: Cowork
#     and Dell's side never see it.
#   * `CARR AI/CLAUDE.md` — by its own header "THE single live source of project
#     instructions", loaded by every CARR-rooted session and reached by Cowork
#     through the rev-7 pointer stub. Its line 15 already reads "Standing rules
#     (short; full text in 00_Context/ai-operating-notes.md)" — a hand-written
#     summary pointing at the file that is retiring, which is precisely the job
#     this block takes over.
# Neither file carries the ORDER 38 doctrine-ownership stamp (single writer, the
# Fable seat); tools/stamp-doctrine-ownership.py's set is three doctrine files
# plus every SKILL.md. So there is no single-writer seat to route this through.
#
# WHAT IT IS NOT: it is not the rules. Nothing here binds. Every line is a
# pointer whose whole content is "a rule about X exists, its text is in the
# store". A gist that grows into the rule is the drift this replaces.

CLAUDE_MD_REL = "CLAUDE.md"
GIST_BLOCK = "rule-gist-index"

# Primary sentence boundaries only. " (" and " — " are deliberately NOT stops:
# a third of the store's statements open with a parenthetical scope note
# ("(team scope, ...)") or an em-dash gloss, and cutting there yields a gist that
# reads as a fragment. MIN_GIST guards the other direction — "Outbound document
# formats:" and "Every CARR session carries the Dr." are both real first stops
# and both say nothing, so a stop before MIN_GIST is skipped and the next one used.
_GIST_STOPS = (". ", "; ", ": ")
MIN_GIST = 40
MAX_GIST = 110


def _gist(statement):
    """The rule's opening claim, cut mechanically. No authoring, no judgement.

    Authored one-liners are the failure mode Joe's ruling names: they get written
    once, the rule changes, and the summary quietly becomes a second, wrong copy.
    This is a pure function of the statement, so it cannot say anything the rule
    does not already say in its first breath.
    """
    s = " ".join((statement or "").split())
    # Parentheticals are dropped BEFORE the cut. Roughly a third of the store's
    # statements open with a scope-and-provenance aside — "(team scope, the
    # boundary stated)", "(personal scope, his request, three mechanisms)" — and a
    # gist that spends its whole length on one says nothing about the rule. Scope
    # and provenance are columns in the store and print in full in the rules file;
    # this line has one job, which is naming the rule.
    stripped = " ".join(re.sub(r"\s*\([^()]*\)", "", s).split())
    if len(stripped) >= MIN_GIST:
        s = stripped
    cut = len(s)
    for stop in _GIST_STOPS:
        i = s.find(stop, MIN_GIST)
        if i != -1:
            cut = min(cut, i)
    g = s[:cut].rstrip(" ,;:.")
    if len(g) > MAX_GIST:
        g = g[:MAX_GIST].rsplit(" ", 1)[0].rstrip(" ,;:.—-") + "…"
    return g


GIST_HEAD = [
    "",
    "**The rule store is the authority. These are pointers, not the rules.**",
    "Each line is a one-line gist of one ACTIVE taught rule. The binding text, its",
    "scope, its provenance and the human's own words live in the store and render in",
    "full to `DNA/compiled-rules-shared.md` (shared) and",
    "`00_Context/compiled-rules-joe.md` (Joe-personal). Read the full text there before",
    "acting on anything below; never quote a gist as the rule. Add a rule with the",
    "`teach` verb and `activate-rule` — never by editing this list, which is",
    "regenerated from the store by `run.sh export --only compiled-rules`.",
    "",
]


def build_rule_gist_index(tmp_path, cur):
    """The generated gist block spliced into CARR AI/CLAUDE.md.

    PARTIAL RENDER. This target does not own its file: CLAUDE.md is hand-authored
    and stays that way, and only the bytes between the two markers are replaced.
    Absent markers are a hard failure, never an append — see exporters/partial.py.
    """
    shared = _fetch_rules(cur, None)
    joe = _fetch_rules(cur, "joe")

    body = list(GIST_HEAD)
    body += [f"*Counts to recite at session start: **{len(shared)} shared**, "
             f"**{len(joe)} Joe-personal**.*", ""]

    for label, rows, where in (
        ("SHARED — binds Joe's sessions and Dell's alike", shared,
         "DNA/compiled-rules-shared.md"),
        ("JOE-PERSONAL — binds Joe's sessions only", joe,
         "00_Context/compiled-rules-joe.md"),
    ):
        body += [f"### {label} · {len(rows)} rule(s) · full text in `{where}`", ""]
        if not rows:
            body += ["*None active.*", ""]
            continue
        for key, heading in RULE_SECTIONS:
            hits = [r for r in rows if _rule_section(r) == key]
            if not hits:
                continue          # the index is a lookup surface; empty rows are noise
            body.append(f"**{heading}**")
            body += [f"- {_gist(r['statement'])}{_rule_ident(r) or ''}" for r in hits]
            body.append("")

    body += [f"*Generated {datetime.now(timezone.utc).isoformat()} · "
             f"{len(shared) + len(joe)} active rule(s) · "
             f"exporters/targets.py build_rule_gist_index*", ""]

    tmp_path.write_text(splice(VAULT / CLAUDE_MD_REL, GIST_BLOCK, body))
    rows = shared + joe
    return len(rows), [[r["statement"], _gist(r["statement"]), r["personal_to"],
                        _rule_section(r)] for r in rows]


# -------- introduction-rules.md (the vendor-politics compile target, ORDER 37) --------
#
# The one file both brains were explicitly instructed to write. It stops being a
# two-writer file here: the RULES come from the store (teach -> activate-rule),
# and the file's own identity prose is carried as a literal below.
#
# WHY THE PROSE IS A LITERAL AND NOT LIFTED FROM A FROZEN COPY: clients-active.md
# lifts its header from repo `frozen-sources/`, but Joe's ruling (a) of
# 2026-08-01 moved every freeze to Drive zips and ORDER 42 is purging
# frozen-sources/ out of git history entirely. An exporter that reads a file
# scheduled for deletion is an exporter that breaks on a date nobody wrote down.
# The prose is version-controlled here instead, which is also where a reader
# looking for "who writes this file now" will actually look.
#
# WHAT IS NOT IN THE STORE, and why: the paragraphs below are the file's
# identity, its data-source pointer and its companion-file pointer. None of them
# parse into a rule's statement/quote/author/status shape, so per the order's
# stop rule they were not forced into rule rows.

INTRO_PROSE_HEAD = [
    "*This file is how Joe and Dell TEACH Claude which vendors can meet. The engine",
    "(`engine.md`) must check every intro candidate AND every event guest list against",
    "these rules before suggesting anything. Every intro Joe approves, rejects, or",
    "corrects becomes a rule here the same session (correction → rule). The vendor world",
    "is highly political: when a pairing is not covered below, ask, do not guess.*",
    "",
    "*Data source: vendor records live in `DNA/Network/vendors.xlsx`. The fields these",
    "rules reference — Category, Vertical, Territory/State, Stage, Links, Rivalry Group —",
    "are columns there.*",
    "",
    "> **Companion file, different axis:** `vendor-intro-timeline.md` covers vendor →",
    "> **CLIENT** intros (which vendor category to introduce to a client at each deal",
    "> stage). THIS file is vendor ↔ vendor (which of Dell's vendors can meet each",
    "> other). Both pull from `vendors.xlsx`.",
    "",
]

# Section order and headings mirror the pre-migration file, because binding force
# differs section to section: a hard block and an unconfirmed default are not the
# same instruction, and a flat list would erase that. Each rule's statement also
# self-labels, so the distinction survives even where a reader sees one rule
# alone, out of the file.
INTRO_SECTIONS = [
    ("hard_blocks",             "Hard blocks (never introduce, never co-invite to the same event)"),
    ("confirmed_pairs",         "Confirmed good pairs (taught and ruled)"),
    ("unconfirmed_defaults",    "Default-plausible pairs — UNCONFIRMED"),
    ("vertical_awareness",      "Vertical awareness"),
    ("cross_category_overlaps", "Cross-category service overlaps"),
    ("delivery_weighted",       "Delivery-weighted matching (the Practice OS steering rule)"),
    ("event_guest_list",        "Event guest-list rule"),
]

# Step 3 of the original loop read "the answer lands here, dated, under
# Confirmed/Blocked" — a hand-edit instruction, and the exact line ORDER 37
# exists to retire. It is rewritten to the teach path. The rest is unchanged.
INTRO_PROSE_LOOP = [
    "## How the teaching loop works",
    "",
    "1. The engine proposes intros (`engine.md` Section 1), already filtered through this file.",
    "2. Joe or Dell approve, reject, or correct. A rejection gets ONE follow-up question:",
    "   \"what's the rule I should learn?\"",
    "3. The answer is captured with the `teach` verb, scope `intro_politics`, in the human's",
    "   verbatim words, and activated with `activate-rule` on his yes. It appears here on the",
    "   next compile. Do not edit this file to record it. Unconfirmed defaults are superseded",
    "   by a confirmed rule, or retired, as they are ruled on.",
    "4. Dell's onboarding (`DNA/Team/twin-system-playbook.md` Phase 3) includes a pairing-rules",
    "   walk — his politics knowledge is deeper in some categories; capture it the same way.",
    "",
]

INTRO_HEADER = [
    "> **GENERATED from the CARR record layer's rule store — do not hand-edit.**",
    "> These rules BIND like the standing rules in `CARR AI/CLAUDE.md`.",
    "> DOMAIN SCOPE `intro_politics`: vendor ↔ vendor compatibility, shared by both partners.",
    "> They are deliberately NOT in `DNA/compiled-rules-shared.md` — that file is the",
    "> always-read core, and this one is read when an intro or a guest list is on the table.",
    ">",
    "> **To add or change a rule, do not edit this file.** Capture it with the `teach` verb",
    "> (the human's verbatim words as `human_quote`, scope `{\"kind\": \"intro_politics\"}`),",
    "> get the human's yes via `activate-rule`, then refresh with",
    "> `run.sh export --only compiled-rules`.",
    "> Only ACTIVE rules appear here — a proposed rule binds nobody by design.",
    "",
]


def _intro_when(r):
    """The date the rule was TAUGHT, not the date it was activated.

    Every rule imported from the pre-migration file was taught weeks before this
    compile target existed, so activated_at would date all nineteen to the day of
    the migration and quietly rewrite the file's own history. The importer stored
    the original date in scope.taught_on; a rule taught live from here on has no
    taught_on and falls back to activated_at, which for it IS the honest date.
    """
    scope = r.get("scope") or {}
    if scope.get("taught_on"):
        return scope["taught_on"]
    if scope.get("import"):
        # An imported rule with no taught_on had no date in the source file.
        # Falling through to activated_at would stamp it with the day of the
        # migration and read as though someone taught it that morning — a
        # fabricated fact, on the surface whose whole job is provenance.
        return "date not recorded in the source file"
    return r["activated_at"].date().isoformat() if r["activated_at"] else "date unrecorded"


def build_rules_intro(tmp_path, cur):
    cur.execute("select * from v_compiled_rules")
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    rows = [r for r in rows if _is_intro(r)]

    lines = ["# Introduction Rules — the taught compatibility layer", ""]
    lines += INTRO_HEADER
    lines += INTRO_PROSE_HEAD

    for key, heading in INTRO_SECTIONS:
        lines += [f"## {heading}", ""]
        hits = [r for r in rows if (r.get("scope") or {}).get("section") == key]
        if not hits:
            # An empty section is PRINTED, never dropped. A missing heading reads
            # as "no such rule category exists"; an empty one reads as "nothing
            # is active here yet", which is the true statement while rules sit
            # proposed and waiting on a human yes.
            lines += ["*No active rule in this section.*", ""]
            continue
        for r in hits:
            line = f"- **{r['statement']}** — taught by {r['taught_by']}, {_intro_when(r)}"
            if r["human_quote"]:
                line += f' ("{r["human_quote"]}")'
            if r["enforcement"] and r["enforcement"] != "prose":
                line += f"  `[{r['enforcement']}]`"
            lines.append(line)
        lines.append("")

    lines += INTRO_PROSE_LOOP

    # The learned-rules log is DERIVED now. It used to be hand-appended, which
    # made it a second place the same fact had to be written and a second place
    # it could go stale. Newest first, one line per rule, carrying the provenance
    # the importer recorded.
    lines += ["## Learned rules log (dated — newest first)", ""]
    if not rows:
        lines += ["*No active rules yet.*", ""]
    # Newest first, with the undated ones LAST: an undated rule is not a recent
    # rule, and a plain reverse sort on the rendered string floats "date not
    # recorded" above every real date.
    def _log_key(x):
        when = _intro_when(x)
        return (0, "") if not when[:1].isdigit() else (1, when)

    for r in sorted(rows, key=_log_key, reverse=True):
        scope = r.get("scope") or {}
        src = scope.get("source") or f"taught by {r['taught_by']}"
        lines.append(f"- {_intro_when(r)} — {r['statement'][:120]}"
                     f"{'…' if len(r['statement']) > 120 else ''}  ·  *{src}*")
    lines.append("")

    from datetime import datetime, timezone
    lines += [f"*Exported: {datetime.now(timezone.utc).isoformat()} · "
              f"{len(rows)} active rule(s)*", ""]
    tmp_path.write_text("\n".join(lines))
    return len(rows), [[r["statement"], r["human_quote"], r["taught_by"],
                        (r.get("scope") or {}).get("section"), r["enforcement"]] for r in rows]


# ---------------- the loop accumulators (one-writer Phase A, ORDER 31) ----------------

LOOP_TARGETS = {
    "open-loops.md": "00_Context/open-loops.md",
    "open-loops-backlog.md": "00_Context/open-loops-backlog.md",
    "action-required.md": "DNA/Team/action-required.md",
    "team-loops.md": "DNA/Team/team-loops.md",
    # ORDER 40. The idea bank is the same machinery with a different kind: its
    # rows are loop_item kind='idea' and its scaffolding is loop_block, and
    # build_loop_file is already generic over both. Nothing here is idea-specific,
    # which is the argument for having put ideas on loop_item at all.
    # UNREGISTERED 2026-08-02, same reason as decision-history.md: registered while
    # RE-REGISTERED 2026-08-03 (Joe: "go ahead and solve decision history and idea
    # bank"). Verified by scratch-render before any vault write, same method as
    # decision-history above: 49 loop_item kind='idea' rows render to 52,621 bytes
    # against the live file's 53,155 — a 534-byte delta with BOTH '## ' sections
    # present and nothing in the live file absent from the render. Unlike the
    # decision target this one is not windowed, so it is a straight replacement.
    # ORDER 40's file swaps are held pending Joe's activation sitting; re-register then.
    "idea-bank.md": "00_Context/idea-bank.md",
}

# NO GENERATED BANNER IS INJECTED INTO THESE FOUR, and that is deliberate.
# Every other generated file opens with one. These four are read by the heartbeat,
# the Monday brief and Dell's sessions as a first act, and each OPENS with a
# doctrine paragraph that IS the rule those readers obey — open-loops.md's marker
# convention, action-required.md's escalation clause. A banner above that prose
# changes the first thing every reader sees and breaks the round-trip diff the
# order's done-test turns on. The do-not-hand-edit warning belongs IN the stored
# prose, added once by a human at the live flip, into the block the human owns.
# Until that flip these renders are staging-only, so nothing sits unlabelled in
# the vault.


def _loop_cell(v):
    """One cell, verbatim.

    NOT _md_cell. That escapes pipes and folds newlines to <br>, which is right
    for a value arriving from a spreadsheet and wrong here: these values came OUT
    of markdown tables carrying their own escapes (team-loops T36 quotes an email
    subject containing an escaped pipe) and one cell legitimately spans two lines
    (T54). Re-escaping would double the backslashes and folding would destroy the
    line break. Both are content changes, on the one surface whose entire test is
    that nothing changed.
    """
    return "" if v is None else str(v)


# Sentinel for "no group emitted yet". None is a real domain (unclassified), so it
# cannot double as the not-started marker without swallowing the first heading.
_NO_DOMAIN_YET = object()


def build_loop_file(rel_path):
    """One builder per file; the render walks the stored blocks in order."""

    def build(tmp_path, cur):
        cur.execute(
            "select seq, block_key, prose_md, header_cols, col_order "
            "from loop_block where rel_path = %s order by seq", (rel_path,))
        blocks = cur.fetchall()
        if not blocks:
            raise ValueError(f"no loop_block rows for {rel_path} — the importer has not run")

        lines, canonical = [], []
        for seq, block_key, prose_md, header_cols, col_order in blocks:
            # A prose-only block is emitted even when EMPTY: the last block of
            # open-loops-backlog.md is exactly that, and it is what carries the
            # file's trailing newline. Dropping it as falsy cost a byte and the
            # round-trip diff caught it.
            if prose_md or block_key is None:
                lines.append(prose_md)
            if block_key is None:
                continue
            # ORDER 0042: grouped by DOMAIN, deals first, unclassified (sort 999)
            # last. Joe: "important transactional and business things get burried by
            # system loops." One flat table was the burial; a table per lane is the fix.
            cur.execute(
                "select row_col_order, number, owner, title, body, since_text, "
                "       unblocks, source_note, closed_text, outcome, "
                "       marker_literal, extra_cells, domain, domain_label "
                "  from v_export_loops "
                " where rel_path = %s and block_key = %s and loop_id is not null "
                " order by domain_sort, render_seq", (rel_path, block_key))
            rows = cur.fetchall()
            # Domain is an OPEN_LOOP axis. team_loop and action_required rows are partner
            # handoffs and cross-brain interrupts — a different question entirely — and
            # their domain is legitimately NULL. Grouping them produced an "Unsorted — no
            # domain set" heading on team-loops.md and action-required.md, which is not a
            # gap to fix but a category error. If no row in this block is classified, the
            # block renders exactly as it did before domains existed.
            group_by_domain = any(r[12] is not None for r in rows)
            current_domain = _NO_DOMAIN_YET
            if not group_by_domain and header_cols:
                lines.append("| " + " | ".join(header_cols) + " |")
                lines.append("|" + "---|" * len(header_cols))
            for r in rows:
                (row_order, number, owner, title, body, since_text, unblocks,
                 source_note, closed_text, outcome, marker_literal, extra,
                 domain, domain_label) = r
                # A heading and a fresh header row per lane. Emitted only when the
                # domain CHANGES, so a block holding one domain looks exactly as it
                # did before — and a block holding none (the closed/DONE tables,
                # which carry no domain) never grows a heading at all.
                if group_by_domain:
                    if domain != current_domain:
                        current_domain = domain
                        if header_cols:
                            lines.append("")
                            lines.append(f"**{domain_label or 'Unsorted — no domain set'}**")
                            lines.append("")
                            lines.append("| " + " | ".join(header_cols) + " |")
                            lines.append("|" + "---|" * len(header_cols))
                vals = {"number": number, "owner": owner, "title": title, "body": body,
                        "since_text": since_text, "unblocks": unblocks,
                        "source_note": source_note, "closed_text": closed_text,
                        "outcome": outcome}
                # The marker literal was split off the item's own text at import;
                # it goes back onto the same cell with the same single space.
                if marker_literal:
                    text_field = "body" if body is not None else "title"
                    vals[text_field] = marker_literal + " " + _loop_cell(vals[text_field])
                order = row_order or col_order
                cells = []
                for name in order:
                    if name.startswith("extra:"):
                        cells.append(_loop_cell((extra or {}).get(name.split(":", 1)[1])))
                    else:
                        cells.append(_loop_cell(vals.get(name)))
                lines.append("| " + " | ".join(cells) + " |")
                canonical.append([number, list(order), cells])

        tmp_path.write_text("\n".join(lines))
        return len(canonical), canonical

    return build


# ---------------- the dossiers (one-writer Phase B, ORDER 36) ----------------

DOSSIER_DIR = "DNA/Clients/prospects"

# The 23 hand-maintained dossiers, enumerated by the RECORD LAYER, not by a
# directory listing: they are exactly the clients carrying `notes_path`
# (`select count(*) from client where notes_path is not null` = 23, verified in
# production 2026-08-01 after Joe's ruling). Each builder re-checks its own
# rel_path against v_export_dossier_subject and refuses to write if it is
# missing, so this list and the database cannot drift apart silently.
#
# WAS 20. Joe's ruling of 2026-08-01 extended the set by three, closing the gap
# this order's first pass flagged — three live, hand-maintained dossiers the
# record did not point at:
#   Ahlborn-FamilyDental-Loxley.md  C-153 Michael Ahlborn
#   Finelli-JosephFinelli.md        C-154 Joseph Finelli
#   VictusDental-Le.md              C-063 Anthony Le — the record existed all
#     along under the practice owner's name, which is why a name-based sweep
#     never matched the file; the supervisor found it and set notes_path.
#
# Two files in the same folder stay OUT, each for its own reason:
#   AltaPointe-enterprise.md — a NATIONAL ACCOUNT, a separate business model with
#     its own lane, and Joe deferred its roster row on 2026-07-22. Excluded by
#     ruling, not by oversight. Do not add it.
#   Beasley-intake.md — an intake, not a dossier (DNA/Clients/INDEX.md:9).
# name -> RENDER MODE. Explicit, stored, never inferred at render time.
#
# Joe's ruling, 2026-08-01, after the tie-break fix showed that only FOUR of the
# 23 dossiers are chronological. The rest are organised by subject, and their
# last section is a trailing appendix (the _template.md layout ends on
# "Requirements snapshot"; the long narratives end on Changelog / Log / History),
# so promoting "the newest" put boilerplate at the top of 18 of 23 renders.
#
#   'chronological' — newest analysis in full, older collapsed to title lines.
#                     Correct where addenda genuinely run down the page.
#   'flat'          — every section at one level, full text, in DOCUMENT ORDER.
#                     No section is promoted, because none of them is "current".
DOSSIER_FILES = {
    "Ahlborn-FamilyDental-Loxley.md": "flat",
    "AmericanFamilyCare.md": "flat",
    "AnointedOT-Sears.md": "flat",
    "BayAreaOralSurgery.md": "flat",
    "Beasley.md": "flat",
    "BlakesEnrickment-Heard.md": "flat",
    "CosmeticDermatology.md": "flat",
    "DeepWaters-Pappas.md": "flat",
    "Finelli-JosephFinelli.md": "flat",
    "FirstCallDPC-Petersen.md": "chronological",
    "GulfCoastPelvicFloor.md": "chronological",
    "HealthcareForKids.md": "flat",
    "Hughes-DentalStartup-SRB.md": "flat",
    "LifeDentalGroup.md": "flat",
    "Lindsey-LighthouseDental.md": "chronological",
    "OceanWounds-Lerner.md": "flat",
    "PCB-Jeremiah-relocation.md": "flat",
    "PremierHealthWellness-RandallMacDonnell.md": "flat",
    "Renalus.md": "flat",
    "SerenityCardiology-Brown.md": "flat",
    "Tyrer-DentalStartup-Moultrie.md": "flat",
    "VictusDental-Le.md": "flat",
    "Weiler-Rejuvime.md": "chronological",
}

# UNLIKE the four loop renders, a dossier DOES carry the generated banner. The
# loop files open with doctrine prose their readers obey, so a banner above it
# changes the first thing every reader sees; a dossier opens with a subject
# heading and its whole point after Phase B is that nobody edits it again
# ("Nobody opens a dossier to edit it again, from any surface, including
# phones" — two-writer-endgame R1). The banner is the instruction that makes
# that true.
def _dossier_banner(mode):
    """The banner has to match what the file below it actually does.

    A flat render promotes nothing, so telling its reader that "newest prints in
    full" would describe a behaviour the page does not have — on the one surface
    whose job is to be trusted without opening the database.
    """
    shape = ("> Newest analysis prints in full, earlier analysis collapses to title +"
             if mode == "chronological" else
             "> Every section renders in full, in the order the source file held them;")
    tail = ("> date + author (the full text is on the record)."
            if mode == "chronological" else
            "> nothing is promoted as 'current' — this dossier is topical, not a log.")
    return [
        "> **GENERATED from the CARR record layer — do not hand-edit.**",
        "> The header below is record fields; the analysis below it is `kind=analysis`",
        "> rows on this subject's timeline.",
        shape,
        tail,
        ">",
        "> **To add analysis, do not edit this file.** Log it against this subject with",
        "> the analysis verb, then refresh with `run.sh export --only dossier`.",
        "",
    ]

# The header fields, in the order the hand-maintained dossiers printed them.
# Only fields that EXIST on the record are listed. The hand files also carried
# Situation / Key angle / Outreach notes / Next step / Requirements snapshot,
# which have no columns on `client`; those arrive as analysis rows at import
# rather than being invented as schema. See the ORDER 36 log's byte-identity note.
DOSSIER_HEADER_FIELDS = [
    ("Client ID", "client_ref"),
    ("Status", "client_status"),
    ("Owner", "owner_label"),
    ("Specialty / Practice type", "specialty_type_label"),
    ("Vertical", "vertical"),
    ("Subtype", "subtype"),
    ("Client type", "client_type"),
    ("Deal type", "deal_type_label"),
    ("Contact", "contact_label"),
    ("Representation (ETL)", "etl_status"),
    ("Lead source", "acquisition_source"),
    ("Lead source detail", "acquisition_detail"),
    ("Last touch", "last_touch"),
]


def _dossier_stamp(occurred_at, author, source=None):
    """Date + author for one analysis row, WITH its confidence. Never fabricated.

    Joe's 2026-08-01 ruling lets a section with no author stamp of its own
    inherit the file's `owner:` frontmatter — which is the difference between a
    reviewable corpus and 84 flagged rows. The cost is that an inherited name
    looks identical to a stamped one unless the render says otherwise, so it
    does. `source` carries the confidence off the imported row:
      import              — the section stamped this author itself
      import_file_stamp   — inherited from the file's owner stamp
      import_unattributed — nobody claimed it; actor is 'system'
    Anything written live through the verb carries 'stated' and needs no
    qualifier: the actor IS the author.
    """
    when = occurred_at.date().isoformat() if occurred_at else "date unrecorded"
    who = author or "author unrecorded"
    src = source or ""
    if src.startswith("import_file_stamp"):
        _, _, stamp = src.partition(":")
        # A compound stamp prints VERBATIM. "Shared, Dell originated" reduced to
        # "dell" would quietly drop both that the client is shared and that Dell
        # brought it in — attribution the partnership tracks on purpose.
        who += (f" (per file stamp \u201c{stamp}\u201d, not this section)" if stamp
                else " (per file stamp, not this section)")
    elif src == "import_unattributed":
        who = "author unrecorded in the source file"
    return f"{when} · {who}"


def build_dossier(rel_path, mode="chronological"):
    """One builder per dossier file: record header, then the analysis stream."""

    def build(tmp_path, cur):
        cur.execute("select * from v_export_dossier_subject where rel_path = %s", (rel_path,))
        cols = [d[0] for d in cur.description]
        subj = cur.fetchall()
        if len(subj) != 1:
            raise ValueError(
                f"{rel_path}: v_export_dossier_subject returned {len(subj)} rows, want 1 — "
                "the client's notes_path moved or 0028 has not been applied")
        s = dict(zip(cols, subj[0]))

        cur.execute(
            "select recency_rank, occurred_at, title, body, owed, author, source "
            "  from v_export_dossier_analysis where rel_path = %s order by recency_rank",
            (rel_path,))
        notes = cur.fetchall()

        lines = [f"# {s['subject_name']}", ""] + _dossier_banner(mode)
        for label, key in DOSSIER_HEADER_FIELDS:
            v = s.get(key)
            if v not in (None, ""):
                lines.append(f"- **{label}:** {v}")
        lines.append("")

        if not notes:
            # Not an error. A dossier whose analysis has not been imported yet
            # renders its header and says so, rather than looking complete.
            lines += ["*No analysis rows on this subject yet — the import for this "
                      "file has not run. Nothing is missing from the record; nothing "
                      "has been written to it either.*", ""]
        elif mode == "flat":
            # Document order. recorded_at was stamped monotonically in file order
            # at import, so reversing recency_rank walks the source file top to
            # bottom. Every section gets the same heading level and its full
            # text: nothing is "current" in a topical dossier.
            lines += ["## Analysis", "",
                      "*Every section of this dossier, in the order the source file held "
                      "them. This dossier is topical rather than a dated log, so no "
                      "section is promoted as the current one.*", ""]
            for _r, occ, ttl, body, owed, auth, src in reversed(notes):
                lines += [f"### {ttl}", f"*{_dossier_stamp(occ, auth, src)}*", ""]
                if body:
                    lines += [body, ""]
                if owed:
                    lines += [f"**Owed:** {owed}", ""]
        else:
            _rank, occurred_at, title, body, owed, author, source = notes[0]
            lines += ["## Analysis — current", "",
                      f"### {title}", f"*{_dossier_stamp(occurred_at, author, source)}*", ""]
            if body:
                lines += [body, ""]
            if owed:
                lines += [f"**Owed:** {owed}", ""]
            if len(notes) > 1:
                lines += ["## Earlier analysis", "",
                          "*Collapsed to title, date and author. The full text lives on the "
                          "record — ask for the subject's timeline.*", ""]
                for _r, occ, ttl, _b, _o, auth, src in notes[1:]:
                    lines.append(f"- **{ttl}** — {_dossier_stamp(occ, auth, src)}")
                lines.append("")

        from datetime import datetime, timezone
        lines += [f"*Exported: {datetime.now(timezone.utc).isoformat()} · "
                  f"{len(notes)} analysis row(s)*", ""]
        tmp_path.write_text("\n".join(lines))

        # Canonical rows for the export checksum: the CONTENT, never the
        # timestamp line (which would make every run look like a change).
        canonical = [[str(s.get(k)) for _l, k in DOSSIER_HEADER_FIELDS]] + [
            [r[0], str(r[1]), r[2], r[3], r[4], r[5]] for r in notes]
        return len(notes), canonical

    return build


# ---------------- decision history (one-writer, ORDER 40) ----------------

DECISION_REL = "00_Context/decision-history.md"

# THE BUDGET THAT REPLACES ORDER 4's MANUAL SPLIT.
#
# decision-history.md has been hand-split three times (Jul 22, Jul 25, Jul 31)
# because a file that every session appends to grows past the 100KB tripwire and
# somebody has to move the old half to an archive file. That whole ritual exists
# only because the file WAS the storage. Now it is a render, so the window is a
# query bound and the archive is the same query with a wider bound: the render
# walks entries newest-first and stops before it would cross this budget.
#
# The consequence, which is the done-test: appending a 193rd entry can never
# require a manual split again. The oldest entry simply falls out of the window
# and stays exactly as reachable as it was — v_decision_entry holds all of them.
# Nothing is deleted, nothing is moved, no second file is created.
#
# Set just under the 100KB tripwire the size sweep enforces.
DECISION_BUDGET_BYTES = 95_000


def build_decision_history(tmp_path, cur):
    """Newest-first, grouped per session (rule 29), windowed by byte budget.

    GROUPING IS HERE AND NOT IN THE VIEW, per R-40a: rule 29 ("one entry per
    session") is doctrine, and doctrine belongs where Joe can change it without
    a migration. v_decision_entry exposes session_key and entry_date and groups
    nothing.
    """
    cur.execute(
        "select entry_date, session_key, title, author, human_quote, agent_rationale, "
        "       quote_absent, provenance "
        "  from v_decision_entry "
        " order by entry_date desc, session_key desc")
    rows = cur.fetchall()

    head = [
        "# Decision History",
        "",
        "> **GENERATED from the CARR record layer — do not hand-edit.**",
        "> Decisions are event rows (`subject_type='decision'`). To record one, use the",
        "> verb; this file is a render of what the record already holds.",
        ">",
        "> **This is a WINDOW, not the whole history.** The render carries the most",
        "> recent entries up to a size budget and stops. Everything older is not",
        "> archived, moved or lost — it is the same query with a wider bound:",
        "> `select * from v_decision_entry where entry_date >= '<date>' order by entry_date desc`.",
        "> ORDER 4's manual 100KB split is retired by this: the window self-limits.",
        "",
    ]

    body, shown, budget = [], 0, DECISION_BUDGET_BYTES - sum(len(l) + 1 for l in head)
    used = 0
    canonical = []
    last_date = None
    for (entry_date, session_key, title, author, quote, rationale,
         quote_absent, provenance) in rows:
        chunk = []
        if entry_date != last_date:
            chunk.append(f"## {entry_date}")
            chunk.append("")
        chunk.append(f"### {title}")
        stamp = f"*{entry_date} · {author}*"
        chunk.append(stamp)
        chunk.append("")
        if quote:
            chunk.append(f"> {quote}")
            chunk.append("")
        if rationale:
            chunk.append(rationale)
            chunk.append("")

        size = sum(len(l) + 1 for l in chunk)
        if used + size > budget and shown:
            break
        body += chunk
        used += size
        shown += 1
        last_date = entry_date
        canonical.append([str(entry_date), session_key, title, author,
                          quote or "", str(quote_absent)])

    omitted = len(rows) - shown
    tail = [
        "---",
        "",
        f"*Window: {shown} of {len(rows)} recorded decisions"
        + (f"; {omitted} older entr{'y' if omitted == 1 else 'ies'} are outside the "
           "window and are read with a wider query, not from a second file."
           if omitted else "; the whole history fits the window.") + "*",
        "",
    ]
    from datetime import datetime, timezone
    tail += [f"*Exported: {datetime.now(timezone.utc).isoformat()} · "
             f"{len(rows)} decision(s) on record*", ""]

    tmp_path.write_text("\n".join(head + body + tail))
    return shown, canonical


TARGETS = {
    "lead-registry.xlsx": (REGISTRY_REL, build_registry),
    # ORDER 40. A render of decision events, windowed by byte budget rather than
    # split by hand — see DECISION_BUDGET_BYTES.
    # RE-REGISTERED 2026-08-03 on Joe's instruction ("go ahead and solve decision
    # history and idea bank"), after the risk that held it was MEASURED rather than
    # assumed. The 2026-08-02 note below was right to stop: bootstrapping blind would
    # have overwritten a live hand-written file. What changed is the verification.
    #
    # The file was rendered to SCRATCH and diffed against the live copy before any
    # vault write. Result: the render is a 95KB WINDOW (DECISION_BUDGET_BYTES) of
    # newest-first, so all 27 of the live file's older sections fall outside it —
    # which looks like catastrophic loss and is not. Every one of those entries is
    # IN the store: 211 decision events spanning 2026-06-29 to 2026-08-03, and the
    # four spot-checked by title ("Life Dental Group onboarded as C-156", "THE RECORD
    # LAYER decided", "practicecre.com registered", the doctorcre ruling) all resolve
    # to real rows. The window drops them from the FILE, not from the record, which
    # is exactly what the byte budget exists to do.
    #
    # The exporter's own .generations/ snapshot preserves the outgoing hand-written
    # file verbatim on the first write, so the pre-render text stays recoverable.
    "decision-history.md": (DECISION_REL, build_decision_history),
    "client-roster.xlsx": (ROSTER_REL, build_roster),
    "vendors.xlsx": (VENDORS_REL, build_vendors),
    "panhandle-team-deals.json": (DEALS_REL, build_deals),
    "clients-active.md": (ACTIVE_REL, build_clients_active),
    "compiled-rules-shared": (RULES_SHARED_REL, build_rules_shared),
    "compiled-rules-joe": (RULES_JOE_REL, build_rules_joe),
    # ORDER 37. Named `compiled-rules-intro` ON PURPOSE: `--only compiled-rules`
    # already prefix-matches, so this file rides the hourly bin/refresh-rules.sh
    # and the nightly chain with no edit to either. A rule taught at 9am reaches
    # the engine at the top of the next hour, same as every other taught rule.
    "compiled-rules-intro": (RULES_INTRO_REL, build_rules_intro),
    # Dell onboarding prep (2026-08-01): Dell's personal compiled rules, in the
    # DNA share. Prefix-matches --only compiled-rules → rides hourly + nightly.
    "compiled-rules-dell": (RULES_DELL_REL, build_rules_dell),
    # PASS 1 (2026-08-03). The one-line gist index inside CARR AI/CLAUDE.md — a
    # PARTIAL render (see build_rule_gist_index and exporters/partial.py); the rest
    # of that file stays hand-authored and byte-identical. Named with the
    # `compiled-rules` prefix on purpose, same reasoning as the intro target: the
    # gist list is a view of the same store, so it must refresh on the same hourly
    # bin/refresh-rules.sh beat. A gist index that lags the rules it indexes is the
    # drift this target exists to prevent.
    "compiled-rules-gist-index": (CLAUDE_MD_REL, build_rule_gist_index),
    # The data dictionary. Structure, the migrations' own COMMENT ON prose, the
    # closed vocabularies and per-column fill counts — no row values (see the
    # module docstring's three rules). Shared tier: both partners read the same
    # record and both ask the same "what does status = 'roster' mean".
    "record-layer-dictionary.md": (DICT_REL, build_dictionary),
    # ORDER 39 (2026-08-01): the two md-ledger renders. hunt-ledger.md is the
    # live flipped file; the reciprocity render lives beside deals.md because
    # deals.md is section-scoped (27 hand-kept deal records stay hand-kept).
    "md-ledger-hunt": (LEDGER_HUNT_REL, build_hunt_ledger),
    "md-ledger-reciprocity": (LEDGER_RECIP_REL, build_reciprocity),
    # #8 (Wave 3, ORDER 25d). Carries a death sentence — see build_router.
    "lead-router-2026-07-13.xlsx": (ROUTER_REL, build_router),
    # #9-#12 (one-writer Phase A, ORDER 31d). `--only loop` refreshes all four,
    # the same prefix-match convenience `--only compiled-rules` relies on.
    **{f"loop-{name}": (rel, build_loop_file(rel)) for name, rel in LOOP_TARGETS.items()},
    # #13-#35 (one-writer Phase B, ORDER 36). `--only dossier` refreshes all 23;
    # `--only dossier-Renalus.md` refreshes exactly one, which is what the
    # file-by-file migration gate in step 8 calls per file.
    **{f"dossier-{name}": (f"{DOSSIER_DIR}/{name}",
                           build_dossier(f"{DOSSIER_DIR}/{name}", mode))
       for name, mode in DOSSIER_FILES.items()},
}
